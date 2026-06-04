"""
Account Reconciliation Tool
- Load large Excel/CSV files (5L+ rows)
- View, Edit, Add, Delete rows
- Triple simultaneous column filter
- Alternating Reconciliation:
    * Sort by date
    * Check first entry — Debit or Credit
    * If Debit first  → find next Credit(s) whose SUM == that Debit amount
    * If Credit first → find next Debit(s)  whose SUM == that Credit amount
    * After each match, flip side and repeat from next unmatched entry
    * Number of entries in a combination is unknown / variable
- Reconcile at All OP Data + Expense Type level separately
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import io
import os
import time
from itertools import combinations

# ─────────────────────────────────────────────
# CONFIG — column name mappings (edit if needed)
# ─────────────────────────────────────────────
COL_DATE      = "Post Date"
COL_DEBIT     = "Debit"
COL_CREDIT    = "Credit"
COL_OP_DATA   = "All OP Data"
COL_EXP_TYPE  = "Expense Type"
COL_NAME      = "Name"

# Expense types that must ONLY reconcile within the same type (strict 1-to-1 category)
STRICT_EXPENSE_TYPES = {"fuel", "insurance", "camera", "prepass", "plates"}

PAGE_SIZE = 500



# ══════════════════════════════════════════════════════════
#  RECONCILIATION ENGINE  (optimised — array-based, int math)
#  Phase 1:   Exact 1:1 matching (debit == credit)
#  Phase 1.5: Consecutive N:1 sum matching (multi-pass)
#  Phase 1.6: Greedy accumulation (all-remaining → single target)
#  Phase 1.7: N:M consecutive matching (window vs window)
#  Phase 2A:  Meet-in-the-middle subset-sum (≤25 entries, any combo)
#  Phase 2B:  Combinatorial fallback (larger groups, combo ≤5)
# ══════════════════════════════════════════════════════════



def reconcile_group(group_df: pd.DataFrame, match_counter_start: int = 1):
    """
    Reconcile a single group using multi-phase matching.

    Phase 1:   Exact 1:1 — debit == credit (across entire group)
    Phase 1.5: Consecutive sum — contiguous sequence of entries on
               one side sums to a single entry on the opposite side

    Returns (df, next_match_counter).
    """
    GROUP_TIME_LIMIT = 30.0  # hard cap: 30 seconds per group
    _group_start = time.perf_counter()

    def _group_timed_out():
        return (time.perf_counter() - _group_start) > GROUP_TIME_LIMIT

    df = group_df.copy().reset_index(drop=True)
    df[COL_DATE] = pd.to_datetime(df[COL_DATE], errors="coerce")
    df = df.sort_values([COL_DATE], kind="mergesort").reset_index(drop=True)

    n = len(df)

    # ── Vectorised pre-extraction ──
    debit_arr  = [0] * n
    credit_arr = [0] * n

    if COL_DEBIT in df.columns:
        dvals = pd.to_numeric(df[COL_DEBIT], errors="coerce").fillna(0.0).values
        for i in range(n):
            if dvals[i] != 0.0:
                debit_arr[i] = int(round(dvals[i] * 100))

    if COL_CREDIT in df.columns:
        cvals = pd.to_numeric(df[COL_CREDIT], errors="coerce").fillna(0.0).values
        for i in range(n):
            if cvals[i] != 0.0:
                credit_arr[i] = int(round(cvals[i] * 100))

    # Result arrays
    matched    = [False] * n
    match_id   = [""] * n
    match_note = [""] * n
    match_type = [""] * n      # "exact" for 1:1, "consecutive" for consecutive-sum
    used       = [False] * n
    match_counter = match_counter_start

    # ──────────────────────────────────────────────────────────────────
    # PHASE 1: Exact 1:1 matches (debit value == credit value)
    # Build a map of credit values → list of indices for O(1) lookup
    # ──────────────────────────────────────────────────────────────────
    credit_map = {}  # paise_value → [list of indices]
    for i in range(n):
        if credit_arr[i] != 0:
            credit_map.setdefault(credit_arr[i], []).append(i)

    for i in range(n):
        if used[i] or debit_arr[i] == 0:
            continue
        d_val = debit_arr[i]
        if d_val in credit_map:
            # Find first unused credit with this value
            clist = credit_map[d_val]
            for ci in clist:
                if not used[ci]:
                    # 1:1 match found
                    mid = f"M{match_counter:04d}"
                    match_counter += 1
                    amount = d_val / 100.0
                    note = f"Debit ₹{amount:,.2f} ↔ Credit ₹{amount:,.2f} (1:1)"

                    for idx in [i, ci]:
                        matched[idx]    = True
                        match_id[idx]   = mid
                        match_note[idx] = note
                        match_type[idx] = "exact"
                        used[idx]       = True
                    break

    # ──────────────────────────────────────────────────────────────────
    # PHASE 1.5: CONSECUTIVE SUM MATCHING  (multi-pass until stable)
    #   Find contiguous/consecutive sequences of debit entries whose
    #   sum equals a single credit entry, and vice versa.
    #   Runs repeatedly: after each round of matches, newly-adjacent
    #   entries may form new consecutive windows.
    # ──────────────────────────────────────────────────────────────────

    def _find_consecutive_sum_matches(source_arr, target_arr, source_label, target_label):
        """Find consecutive sequences in source_arr that sum to a target in target_arr.
        Returns number of new matches found (0 means done).
        """
        nonlocal match_counter
        matches_found = 0

        # Collect unused source indices (in order)
        src_indices = [i for i in range(n) if source_arr[i] != 0 and not used[i]]
        if not src_indices:
            return 0

        # Collect unused target values → indices for lookup
        tgt_map = {}  # paise_value → [list of indices]
        for i in range(n):
            if target_arr[i] != 0 and not used[i]:
                tgt_map.setdefault(target_arr[i], []).append(i)
        if not tgt_map:
            return 0

        for l in range(len(src_indices)):
            if used[src_indices[l]]:
                continue
            if _group_timed_out():
                return matches_found
            running_sum = 0
            for r in range(l, len(src_indices)):
                if r > l:
                    gap_has_source = False
                    for g in range(src_indices[r-1] + 1, src_indices[r]):
                        if source_arr[g] != 0 and not used[g]:
                            gap_has_source = True
                            break
                    if gap_has_source:
                        break

                idx_r = src_indices[r]
                if used[idx_r]:
                    break
                running_sum += source_arr[idx_r]
                count = r - l + 1

                if count < 2:
                    continue

                if running_sum in tgt_map:
                    tgt_list = tgt_map[running_sum]
                    matched_tgt = None
                    for ti in tgt_list:
                        if not used[ti]:
                            matched_tgt = ti
                            break
                    if matched_tgt is not None:
                        mid = f"M{match_counter:04d}"
                        match_counter += 1
                        amount = running_sum / 100.0
                        consec_indices = [src_indices[k] for k in range(l, r + 1)]
                        note = (f"{count} consecutive {source_label} entries "
                                f"₹{amount:,.2f} ↔ {target_label} ₹{amount:,.2f} "
                                f"(consecutive sum)")

                        all_indices = consec_indices + [matched_tgt]
                        for idx in all_indices:
                            matched[idx]    = True
                            match_id[idx]   = mid
                            match_note[idx] = note
                            match_type[idx] = "consecutive"
                            used[idx]       = True

                        tgt_list.remove(matched_tgt)
                        if not tgt_list:
                            del tgt_map[running_sum]

                        matches_found += 1
                        break
        return matches_found

    # Run consecutive N:1 matching in multiple passes until stable
    for _pass in range(20):
        if _group_timed_out():
            break
        found = 0
        found += _find_consecutive_sum_matches(debit_arr, credit_arr, "debit", "Credit")
        found += _find_consecutive_sum_matches(credit_arr, debit_arr, "credit", "Debit")
        if found == 0:
            break

    # ──────────────────────────────────────────────────────────────────
    # PHASE 1.6: GREEDY ACCUMULATION  (all-remaining → single target)
    #   For each unmatched entry on one side, check if ALL remaining
    #   unmatched entries on the other side sum to it.  Catches large
    #   N:1 patterns (e.g. 14 debits → 1 credit) in O(n).
    # ──────────────────────────────────────────────────────────────────

    def _greedy_all_remaining(source_arr, target_arr, source_label, target_label):
        """Check if all remaining unused source entries sum to a single target."""
        nonlocal match_counter
        matches_found = 0

        src_indices = [i for i in range(n) if source_arr[i] != 0 and not used[i]]
        if len(src_indices) < 2:
            return 0
        src_total = sum(source_arr[i] for i in src_indices)

        for i in range(n):
            if target_arr[i] != 0 and not used[i] and target_arr[i] == src_total:
                mid = f"M{match_counter:04d}"
                match_counter += 1
                amount = src_total / 100.0
                note = (f"{len(src_indices)} {source_label} entries "
                        f"₹{amount:,.2f} ↔ {target_label} ₹{amount:,.2f} "
                        f"(sum match)")
                all_matched = src_indices + [i]
                for idx in all_matched:
                    matched[idx]    = True
                    match_id[idx]   = mid
                    match_note[idx] = note
                    match_type[idx] = "sum"
                    used[idx]       = True
                matches_found += 1
                break  # all source entries are now used
        return matches_found

    for _pass in range(10):
        if _group_timed_out():
            break
        found = 0
        found += _greedy_all_remaining(debit_arr, credit_arr, "debit", "Credit")
        found += _greedy_all_remaining(credit_arr, debit_arr, "credit", "Debit")
        if found == 0:
            break

    # ──────────────────────────────────────────────────────────────────
    # PHASE 1.7: N:M CONSECUTIVE MATCHING
    #   Contiguous window of debits whose sum == contiguous window of
    #   credits.  Both windows must have ≥ 2 entries (N:1 already done).
    # ──────────────────────────────────────────────────────────────────

    MAX_WINDOWS = 50_000  # cap to prevent memory/time blowup

    def _get_consecutive_windows(arr, min_size=1, max_size=15):
        """Return list of (indices_list, window_sum) for consecutive unused entries."""
        windows = []
        idxs = [i for i in range(n) if arr[i] != 0 and not used[i]]
        if len(idxs) < min_size:
            return windows
        for l in range(len(idxs)):
            if _group_timed_out() or len(windows) >= MAX_WINDOWS:
                break
            if used[idxs[l]]:
                continue
            running = 0
            for r in range(l, min(l + max_size, len(idxs))):
                if r > l:
                    gap = False
                    for g in range(idxs[r-1] + 1, idxs[r]):
                        if arr[g] != 0 and not used[g]:
                            gap = True
                            break
                    if gap:
                        break
                if used[idxs[r]]:
                    break
                running += arr[idxs[r]]
                sz = r - l + 1
                if sz >= min_size:
                    windows.append(([idxs[k] for k in range(l, r + 1)], running))
        return windows

    # Build debit and credit windows, match by sum
    if not _group_timed_out():
        d_windows = _get_consecutive_windows(debit_arr, min_size=2, max_size=10)
        c_win_map = {}  # sum → list of index-lists
        for c_idx_list, c_sum in _get_consecutive_windows(credit_arr, min_size=2, max_size=10):
            c_win_map.setdefault(c_sum, []).append(c_idx_list)

        for d_idx_list, d_sum in d_windows:
            if _group_timed_out():
                break
            if any(used[i] for i in d_idx_list):
                continue
            if d_sum not in c_win_map:
                continue
            for c_idx_list in c_win_map[d_sum]:
                if any(used[i] for i in c_idx_list):
                    continue
                # N:M match found
                mid = f"M{match_counter:04d}"
                match_counter += 1
                amount = d_sum / 100.0
                note = (f"{len(d_idx_list)} debit + {len(c_idx_list)} credit entries "
                        f"₹{amount:,.2f} (N:M consecutive)")
                for idx in d_idx_list + c_idx_list:
                    matched[idx]    = True
                    match_id[idx]   = mid
                    match_note[idx] = note
                    match_type[idx] = "consecutive"
                    used[idx]       = True
                break

    # ──────────────────────────────────────────────────────────────────
    # PHASE 2: GLOBAL (NON-CONSECUTIVE) SUM MATCHING
    #   Strategy A — Meet-in-the-middle subset-sum for ≤ MITM_MAX
    #     source entries: O(2^(n/2)), finds ANY subset that sums to
    #     a target.  Handles 6–25 entry combos that combinatorial
    #     approach cannot reach.
    #   Strategy B — Combinatorial fallback for larger groups.
    # ──────────────────────────────────────────────────────────────────

    MITM_MAX = 25              # max source entries for meet-in-the-middle
    MAX_COMBO = 5              # fallback combo size for large groups
    MAX_SRC_BIG = 80
    MAX_ITERATIONS = 500_000
    TIME_LIMIT = 10.0

    _phase2_start = time.perf_counter()  # separate timer for Phase 2

    # ── Strategy A: Meet-in-the-middle subset-sum ─────────────────────

    MITM_TIME_LIMIT = 5.0   # seconds per MITM call

    def _mitm_find_subset(entries, target):
        """Find a subset (size ≥ 2) of entries that sums exactly to target.
        entries: list of (index, value).  Returns list of indices or None.
        Uses meet-in-the-middle: O(2^(n/2)) time and space.
        """
        ne = len(entries)
        if ne < 2:
            return None
        _mitm_start = time.perf_counter()
        mid_pt = ne // 2
        left, right = entries[:mid_pt], entries[mid_pt:]

        # All non-empty subsets of left half → {sum: (indices, count)}
        left_sums = {}  # sum → (indices_list, subset_size)
        for mask in range(1, 1 << len(left)):
            if mask % 10_000 == 0 and (time.perf_counter() - _mitm_start) > MITM_TIME_LIMIT:
                return None  # bail out
            s, idxs = 0, []
            for i in range(len(left)):
                if mask & (1 << i):
                    s += left[i][1]
                    idxs.append(left[i][0])
            if s <= target and s not in left_sums:
                left_sums[s] = idxs

        # Left-only hit
        if target in left_sums and len(left_sums[target]) >= 2:
            return left_sums[target]

        # Right subsets, check complement in left
        for mask in range(1, 1 << len(right)):
            if mask % 10_000 == 0 and (time.perf_counter() - _mitm_start) > MITM_TIME_LIMIT:
                return None  # bail out
            s, idxs = 0, []
            for i in range(len(right)):
                if mask & (1 << i):
                    s += right[i][1]
                    idxs.append(right[i][0])
            if s > target:
                continue
            if s == target and len(idxs) >= 2:
                return idxs
            complement = target - s
            if complement in left_sums:
                combined = left_sums[complement] + idxs
                if len(combined) >= 2:
                    return combined
        return None

    def _find_global_mitm(source_arr, target_arr, source_label, target_label):
        """Meet-in-the-middle global sum matching for small groups."""
        nonlocal match_counter
        src = [(i, source_arr[i]) for i in range(n)
               if source_arr[i] != 0 and not used[i]]
        if len(src) < 2 or len(src) > MITM_MAX:
            return False

        targets = [(i, target_arr[i]) for i in range(n)
                   if target_arr[i] != 0 and not used[i]]
        if not targets:
            return False

        found_any = False
        for tgt_idx, tgt_val in targets:
            if used[tgt_idx]:
                continue
            src = [(i, v) for i, v in src if not used[i]]
            if len(src) < 2:
                break
            result = _mitm_find_subset(src, tgt_val)
            if result is not None:
                mid = f"M{match_counter:04d}"
                match_counter += 1
                amount = tgt_val / 100.0
                note = (f"{len(result)} {source_label} entries "
                        f"₹{amount:,.2f} ↔ {target_label} ₹{amount:,.2f} "
                        f"(sum match)")
                for idx in result + [tgt_idx]:
                    matched[idx]    = True
                    match_id[idx]   = mid
                    match_note[idx] = note
                    match_type[idx] = "sum"
                    used[idx]       = True
                found_any = True
        return found_any

    # Run MITM in passes until stable
    for _mitm_pass in range(10):
        if _group_timed_out():
            break
        found_mitm = False
        d_rem = sum(1 for i in range(n) if debit_arr[i] != 0 and not used[i])
        c_rem = sum(1 for i in range(n) if credit_arr[i] != 0 and not used[i])
        if 2 <= d_rem <= MITM_MAX:
            found_mitm |= _find_global_mitm(debit_arr, credit_arr, "debit", "Credit")
        if 2 <= c_rem <= MITM_MAX:
            found_mitm |= _find_global_mitm(credit_arr, debit_arr, "credit", "Debit")
        if not found_mitm:
            break

    # ── Strategy B: Combinatorial fallback for remaining/larger groups ─

    def _find_global_sum_matches(source_arr, target_arr, source_label, target_label):
        nonlocal match_counter
        if _group_timed_out():
            return
        _phase2b_start = time.perf_counter()  # own timer for this call
        src = [(i, source_arr[i]) for i in range(n) if source_arr[i] != 0 and not used[i]]
        if len(src) < 2:
            return

        tgt_map = {}
        for i in range(n):
            if target_arr[i] != 0 and not used[i]:
                tgt_map.setdefault(target_arr[i], []).append(i)
        if not tgt_map:
            return

        effective_max = MAX_COMBO
        if len(src) > MAX_SRC_BIG:
            effective_max = min(effective_max, 3)

        iteration_count = 0

        for combo_size in range(2, effective_max + 1):
            src = [(i, v) for i, v in src if not used[i]]
            if len(src) < combo_size:
                break

            for combo in combinations(src, combo_size):
                iteration_count += 1
                if iteration_count > MAX_ITERATIONS:
                    return
                if iteration_count % 10_000 == 0:
                    if time.perf_counter() - _phase2b_start > TIME_LIMIT:
                        return
                    if _group_timed_out():
                        return

                if any(used[idx] for idx, _ in combo):
                    continue
                total = sum(v for _, v in combo)
                if total not in tgt_map:
                    continue
                tgt_list = tgt_map[total]
                matched_tgt = None
                for ti in tgt_list:
                    if not used[ti]:
                        matched_tgt = ti
                        break
                if matched_tgt is None:
                    continue

                mid = f"M{match_counter:04d}"
                match_counter += 1
                amount = total / 100.0
                combo_indices = [idx for idx, _ in combo]
                note = (f"{combo_size} {source_label} entries "
                        f"₹{amount:,.2f} ↔ {target_label} ₹{amount:,.2f} "
                        f"(sum match)")

                for idx in combo_indices + [matched_tgt]:
                    matched[idx]    = True
                    match_id[idx]   = mid
                    match_note[idx] = note
                    match_type[idx] = "sum"
                    used[idx]       = True

                tgt_list.remove(matched_tgt)
                if not tgt_list:
                    del tgt_map[total]

    if not _group_timed_out():
        _find_global_sum_matches(debit_arr, credit_arr, "debit", "Credit")
    if not _group_timed_out():
        _find_global_sum_matches(credit_arr, debit_arr, "credit", "Debit")

    # Write results back in bulk
    df["_matched"]    = matched
    df["_match_id"]   = match_id
    df["_match_note"] = match_note
    df["_match_type"] = match_type

    # Assign status
    statuses = []
    for i in range(n):
        if matched[i] and match_type[i] == "sum":
            statuses.append("🧮 Sum Match")
        elif matched[i] and match_type[i] == "consecutive":
            statuses.append("📊 Consecutive Match")
        elif matched[i]:
            statuses.append("✅ Matched")
        else:
            statuses.append("❌ Unmatched")
    df["_status"] = statuses
    return df, match_counter


def _row_group_key(row, has_op_data: bool, has_name: bool, has_exp_type: bool) -> tuple:
    """
    Build a (entity_key, expense_group) tuple for a row.

    Entity key priority:
      1. All OP Data  (if non-blank)
      2. Name         (fallback when All OP Data is blank/null)
      3. '__unknown__'

    Expense group:
      - Strict types (fuel, insurance, camera, prepass, plates) → exact type name
      - All other types → '__other__'  (can cross-match freely)
    """
    # ── Entity ──────────────────────────────────────────────────────────────
    entity = "__unknown__"
    if has_op_data:
        op = row[COL_OP_DATA]
        if pd.notna(op) and str(op).strip() not in ("", "nan"):
            entity = str(op).strip().lower()
    if entity == "__unknown__" and has_name:
        nm = row[COL_NAME]
        if pd.notna(nm) and str(nm).strip() not in ("", "nan"):
            entity = str(nm).strip().lower()

    # ── Expense group ────────────────────────────────────────────────────────
    exp_group = "__other__"
    if has_exp_type:
        raw = row[COL_EXP_TYPE]
        exp = str(raw).strip().lower() if pd.notna(raw) and str(raw).strip() not in ("", "nan") else ""
        if exp in STRICT_EXPENSE_TYPES:
            exp_group = exp          # strict: keep exact type so they never cross-match
        # else: leave as '__other__' — all non-strict types pool together

    return (entity, exp_group)


def run_reconciliation(df: pd.DataFrame, progress_cb=None) -> pd.DataFrame:
    """
    New grouping logic:
      • Entity = All OP Data  (primary)  →  Name  (fallback if OP Data is blank)
      • Expense group:
          - fuel / insurance / camera / prepass / plates  →  strict (same-type only)
          - everything else (advance, repair, eps, paycheck, void, …)  → '__other__'
            these pool together and can cross-match freely within the same entity

    progress_cb: optional callable(msg: str) for UI progress updates.
    """
    has_op_data  = COL_OP_DATA  in df.columns
    has_name     = COL_NAME     in df.columns
    has_exp_type = COL_EXP_TYPE in df.columns

    if not has_op_data and not has_name:
        if progress_cb:
            progress_cb("Reconciling (single group)…")
        result, _ = reconcile_group(df)
        return result

    df = df.copy()

    # ── Vectorised group-key construction (replaces slow df.apply) ────────
    # Entity: primary = All OP Data, fallback = Name
    entity = pd.Series("__unknown__", index=df.index)
    if has_op_data:
        op = df[COL_OP_DATA].fillna("").astype(str).str.strip().str.lower()
        mask_op = ~op.isin(["", "nan"])
        entity = entity.where(~mask_op, op)
    if has_name:
        nm = df[COL_NAME].fillna("").astype(str).str.strip().str.lower()
        mask_nm = (entity == "__unknown__") & ~nm.isin(["", "nan"])
        entity = entity.where(~mask_nm, nm)

    # Expense group: strict types stay as-is, everything else → '__other__'
    exp_group = pd.Series("__other__", index=df.index)
    if has_exp_type:
        exp_raw = df[COL_EXP_TYPE].fillna("").astype(str).str.strip().str.lower()
        is_strict = exp_raw.isin(STRICT_EXPENSE_TYPES)
        exp_group = exp_group.where(~is_strict, exp_raw)

    df["_grp"] = list(zip(entity, exp_group))

    groups = list(df.groupby("_grp", dropna=False, sort=False))
    total_groups = len(groups)
    results = []
    global_counter = 1
    failed_groups = 0
    for idx, (_key, grp) in enumerate(groups, 1):
        if progress_cb and (idx % 5 == 0 or idx == 1):
            progress_cb(f"Processing group {idx}/{total_groups} ({len(grp)} rows)…")
        try:
            reconciled, global_counter = reconcile_group(
                grp.drop(columns=["_grp"]), match_counter_start=global_counter
            )
            results.append(reconciled)
        except Exception:
            # If a single group fails, keep its rows unmatched rather than
            # crashing the entire reconciliation.
            failed_grp = grp.drop(columns=["_grp"]).copy()
            failed_grp["_matched"]    = False
            failed_grp["_match_id"]   = ""
            failed_grp["_match_note"] = ""
            failed_grp["_match_type"] = ""
            failed_grp["_status"]     = "❌ Unmatched"
            results.append(failed_grp)
            failed_groups += 1

    result = pd.concat(results, ignore_index=True)

    if progress_cb and failed_groups > 0:
        progress_cb(f"⚠ {failed_groups} group(s) could not be reconciled — marked Unmatched")

    # ── POST-PROCESS: relabel cross-type matches as CM001, CM002, … ──────
    if has_exp_type and COL_EXP_TYPE in result.columns:
        if progress_cb:
            progress_cb("Detecting cross-type matches…")
        cm_counter = 1
        # Vectorised: group by match_id quickly
        matched_mask = result["_match_id"] != ""
        if matched_mask.any():
            id_groups = result.loc[matched_mask].groupby("_match_id", sort=False)
            for mid, grp_rows in id_groups:
                exp_types = set(
                    str(t).strip().lower()
                    for t in grp_rows[COL_EXP_TYPE].dropna()
                    if str(t).strip().lower() not in ("", "nan")
                )
                if len(exp_types) > 1:
                    mask = result["_match_id"] == mid
                    new_id = f"CM{cm_counter:03d}"
                    cm_counter += 1
                    cross_types = ", ".join(sorted(exp_types))
                    old_note = grp_rows["_match_note"].iloc[0] if "_match_note" in grp_rows.columns else ""
                    new_note = f"{old_note} [Cross: {cross_types}]"
                    result.loc[mask, "_match_id"]   = new_id
                    result.loc[mask, "_match_note"] = new_note
                    result.loc[mask, "_status"]     = "🔄 Cross Matched"

    return result


# ══════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════
#  STREAMLIT UI  —  fixed & complete
# ══════════════════════════════════════════════════════════

# ── Excel export (returns bytes) ──────────────────────────

def _export_styled_excel(df_export: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    header_fill    = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_font    = Font(name="Segoe UI", bold=True, color="E94560", size=10)
    matched_fill   = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    matched_font   = Font(name="Segoe UI", color="1A6B3C", size=10)
    consec_fill    = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    consec_font    = Font(name="Segoe UI", color="1A4D8E", size=10)
    sum_fill       = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
    sum_font       = Font(name="Segoe UI", color="6C3483", size=10)
    cross_fill     = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
    cross_font     = Font(name="Segoe UI", color="7D6608", size=10)
    unmatched_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    unmatched_font = Font(name="Segoe UI", color="922B21", size=10)
    default_font   = Font(name="Segoe UI", size=10)
    thin_border    = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    cols = list(df_export.columns)
    status_col_idx = next((i for i, c in enumerate(cols) if c == "_status"), None)

    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, (_, row) in enumerate(df_export.iterrows(), 2):
        status = ""
        if status_col_idx is not None:
            try:
                val = row.iloc[status_col_idx]
                status = str(val) if pd.notna(val) else ""
            except Exception:
                status = ""

        if "Cross" in status:
            row_fill, row_font = cross_fill, cross_font
        elif "Consecutive" in status:
            row_fill, row_font = consec_fill, consec_font
        elif "Sum Match" in status:
            row_fill, row_font = sum_fill, sum_font
        elif "Matched" in status and "Unmatched" not in status:
            row_fill, row_font = matched_fill, matched_font
        elif "Unmatched" in status:
            row_fill, row_font = unmatched_fill, unmatched_font
        else:
            row_fill, row_font = None, default_font

        for col_idx, col_name in enumerate(cols, 1):
            try:
                value = row[col_name]
                cell = ws.cell(row=row_idx, column=col_idx,
                               value=str(value) if pd.notna(value) else "")
            except Exception:
                cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.font = row_font
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

    for col_idx, col_name in enumerate(cols, 1):
        max_len = len(str(col_name))
        for r in range(2, min(52, ws.max_row + 1)):
            cell_val = ws.cell(row=r, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[
            ws.cell(row=1, column=col_idx).column_letter
        ].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Column-mapping helper ─────────────────────────────────

def _best_match(hints, current_default, file_cols):
    if current_default in file_cols:
        return current_default
    lower_cols = {c.lower().strip(): c for c in file_cols}
    for hint in hints:
        if hint.lower() in lower_cols:
            return lower_cols[hint.lower()]
    for hint in hints:
        for lc, orig in lower_cols.items():
            if hint.lower() in lc or lc in hint.lower():
                return orig
    return file_cols[0] if file_cols else ""


# ── Row highlighting ──────────────────────────────────────

def _row_highlight(status: str) -> str:
    if "Cross" in status:
        return "background-color: #fef9c3"
    elif "Consecutive" in status:
        return "background-color: #dbeafe"
    elif "Sum Match" in status:
        return "background-color: #ede9fe"
    elif "Matched" in status and "Unmatched" not in status:
        return "background-color: #dcfce7"
    elif "Unmatched" in status:
        return "background-color: #fee2e2"
    return ""


# ─────────────────────────────────────────────────────────
#  SESSION-STATE HELPERS
# ─────────────────────────────────────────────────────────

def _init_state():
    defaults = {
        "df_original":       None,
        "df_display":        None,
        "hidden_columns":    set(),
        "file_loaded":       False,
        "mapping_confirmed": False,
        "_header_selected":  False,
        "_raw_preview":      None,
        "_raw_bytes":        None,   # FIX: store raw bytes so seek works after reruns
        "_raw_filename":     None,
        "_full_raw_df":      None,
        "header_row":        0,
        "status_msg":        "",
        "page":              0,
        "sort_col":          None,
        "sort_asc":          True,
        # Column mappings stored in session state so they survive reruns
        "col_date":          "Post Date",
        "col_debit":         "Debit",
        "col_credit":        "Credit",
        "col_op_data":       "All OP Data",
        "col_exp_type":      "Expense Type",
        "col_name":          "Name",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _get_cols():
    """Return the current column name mapping from session state."""
    return {
        "date":     st.session_state.col_date,
        "debit":    st.session_state.col_debit,
        "credit":   st.session_state.col_credit,
        "op_data":  st.session_state.col_op_data,
        "exp_type": st.session_state.col_exp_type,
        "name":     st.session_state.col_name,
    }


# ─────────────────────────────────────────────────────────
#  MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="AccountSync · Reconciliation",
        page_icon="◈",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.markdown("""
    <style>
    .main .block-container { padding-top: 1rem; }
    .stButton>button { border-radius: 8px; font-weight: 600; }
    .stat-card { background:#fff; border:1px solid #e2e8f0; border-radius:10px;
                 padding:10px 16px; text-align:center; }
    .stat-label { font-size:11px; color:#64748b; font-weight:700; letter-spacing:.05em; }
    .stat-value { font-size:22px; font-weight:800; margin-top:2px; }
    </style>
    """, unsafe_allow_html=True)

    _init_state()

    # ── Header ────────────────────────────────────────────
    st.markdown("""
    <div style="display:flex;align-items:center;gap:12px;margin-bottom:.5rem;">
      <span style="font-size:28px;color:#3b82f6;">◈</span>
      <div>
        <span style="font-size:20px;font-weight:800;color:#1e293b;">AccountSync</span>
        <span style="font-size:12px;color:#64748b;margin-left:8px;">Reconciliation Tool</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Sidebar ───────────────────────────────────────────
    with st.sidebar:
        _sidebar_load()
        st.divider()
        _sidebar_actions()
        st.divider()
        _sidebar_export()
        st.divider()
        _sidebar_edit()

    # ── Status message ────────────────────────────────────
    if st.session_state.status_msg:
        st.info(st.session_state.status_msg)

    # ── Wizard: header row + column mapping ───────────────
    if st.session_state.file_loaded and not st.session_state.mapping_confirmed:
        _wizard()
        return

    # ── Empty state ───────────────────────────────────────
    if st.session_state.df_original is None:
        st.markdown("""
        <div style="text-align:center;padding:60px 0;color:#94a3b8;">
          <div style="font-size:64px;">📊</div>
          <div style="font-size:20px;font-weight:700;margin-top:12px;">
            Upload a file to get started
          </div>
          <div style="font-size:14px;margin-top:6px;">
            Supports Excel (.xlsx / .xls) and CSV files
          </div>
        </div>
        """, unsafe_allow_html=True)
        return

    # ── Main content ──────────────────────────────────────
    _render_stats_ribbon()
    _render_filters()
    _render_table()
    _render_show_match_section()


# ─────────────────────────────────────────────────────────
#  SIDEBAR SECTIONS
# ─────────────────────────────────────────────────────────

def _sidebar_load():
    st.markdown("### 📂 Load File")

    uploaded = st.file_uploader(
        "Upload Excel or CSV",
        type=["xlsx", "xls", "csv"],
        key="file_uploader",
    )

    # FIX: detect a *new* upload by comparing filename/size, not a boolean flag
    if uploaded is not None:
        new_sig = f"{uploaded.name}_{uploaded.size}"
        if st.session_state.get("_upload_sig") != new_sig:
            st.session_state["_upload_sig"]      = new_sig
            st.session_state["_raw_bytes"]        = uploaded.read()   # store bytes
            st.session_state["_raw_filename"]     = uploaded.name
            st.session_state["file_loaded"]       = True
            st.session_state["mapping_confirmed"] = False
            st.session_state["_header_selected"]  = False
            st.session_state["df_original"]       = None
            st.session_state["df_display"]        = None
            st.session_state["status_msg"]        = ""
            # Build 20-row preview from bytes
            try:
                buf = io.BytesIO(st.session_state["_raw_bytes"])
                if uploaded.name.lower().endswith((".xlsx", ".xls")):
                    prev = pd.read_excel(buf, header=None, nrows=20, dtype=str)
                else:
                    buf.seek(0)
                    prev = pd.read_csv(buf, header=None, nrows=20, dtype=str,
                                       encoding="utf-8-sig")
                st.session_state["_raw_preview"] = prev
            except Exception as e:
                st.error(f"Could not read file: {e}")
                st.session_state["file_loaded"] = False

    if st.session_state.df_original is not None:
        st.success(f"✅ {len(st.session_state.df_original):,} rows loaded")


def _sidebar_actions():
    st.markdown("### ⚙️ Actions")
    has_data = st.session_state.df_original is not None

    c1, c2 = st.columns(2)
    with c1:
        if st.button("🔁 Reconcile", use_container_width=True, disabled=not has_data):
            _run_reconciliation()
    with c2:
        if st.button("🔄 Reset View", use_container_width=True, disabled=not has_data):
            st.session_state.df_display = st.session_state.df_original.copy()
            st.session_state.page = 0
            st.session_state.status_msg = ""
            st.rerun()


def _sidebar_export():
    st.markdown("### 🔽 Export")
    df = st.session_state.df_display
    if df is None:
        st.caption("Load a file first.")
        return

    export_cols = [c for c in df.columns if c != "_matched"]
    df_export   = df[export_cols]
    fmt = st.radio("Format", ["Excel (.xlsx)", "CSV (.csv)"], key="export_fmt",
                   horizontal=True)

    if fmt == "Excel (.xlsx)":
        try:
            data = _export_styled_excel(df_export)
            st.download_button(
                "⬇ Download Excel", data,
                "reconciliation_export.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except ImportError:
            st.warning("openpyxl not installed — exporting as CSV.")
            st.download_button("⬇ Download CSV",
                               df_export.to_csv(index=False).encode(),
                               "reconciliation_export.csv", "text/csv",
                               use_container_width=True)
    else:
        st.download_button("⬇ Download CSV",
                           df_export.to_csv(index=False).encode(),
                           "reconciliation_export.csv", "text/csv",
                           use_container_width=True)


def _sidebar_edit():
    st.markdown("### ✏️ Edit Data")
    if st.session_state.df_original is None:
        st.caption("Load a file first.")
        return

    edit_mode = st.selectbox("Action", [
        "— select —",
        "Add Row",
        "Edit Row by Index",
        "Delete Row by Index",
        "Bulk Edit Column",
        "Add Computed Column",
        "Hide / Show Columns",
        "Delete Columns",
    ], key="edit_mode")

    df = st.session_state.df_original

    if edit_mode == "Add Row":
        _edit_add_row(df)
    elif edit_mode == "Edit Row by Index":
        _edit_edit_row(df)
    elif edit_mode == "Delete Row by Index":
        _edit_delete_row(df)
    elif edit_mode == "Bulk Edit Column":
        _edit_bulk(df)
    elif edit_mode == "Add Computed Column":
        _edit_add_column(df)
    elif edit_mode == "Hide / Show Columns":
        _edit_hide_show(df)
    elif edit_mode == "Delete Columns":
        _edit_delete_columns(df)


# ─────────────────────────────────────────────────────────
#  UPLOAD WIZARD  (header row → column mapping)
# ─────────────────────────────────────────────────────────

def _wizard():
    raw_preview = st.session_state.get("_raw_preview")
    if raw_preview is None:
        return

    # ── Step 1: pick header row ───────────────────────────
    if not st.session_state.get("_header_selected", False):
        st.markdown("### 📋 Step 1 — Select Header Row")
        st.caption("The first 20 rows of your file are shown. Choose which row contains the column headers.")

        prev = raw_preview.copy()
        prev.index.name = "Row #"
        st.dataframe(prev, use_container_width=True, height=320)

        hr = st.number_input(
            "Header row number (0 = first row)",
            min_value=0, max_value=max(0, len(raw_preview) - 1),
            value=0, step=1, key="wizard_header_row",
        )

        if st.button("✅ Use this row as header", type="primary"):
            st.session_state["header_row"]      = int(hr)
            st.session_state["_header_selected"] = True
            _load_full_file_from_bytes(int(hr))
            st.rerun()
        return

    # ── Step 2: map columns ───────────────────────────────
    df = st.session_state.get("_full_raw_df")
    if df is None:
        st.error("Full file not loaded yet. Please go back and re-select the header row.")
        if st.button("↩ Back"):
            st.session_state["_header_selected"] = False
            st.rerun()
        return

    st.markdown("### 🗂 Step 2 — Map Columns")
    st.caption("Match each required field to the correct column in your file.")

    file_cols = list(df.columns)
    options   = ["(None — skip)"] + file_cols

    mapping_fields = [
        ("Date Column",         st.session_state.col_date,     ["date", "post date", "posting date", "txn date"]),
        ("Debit Column",        st.session_state.col_debit,    ["debit", "dr", "debit amount"]),
        ("Credit Column",       st.session_state.col_credit,   ["credit", "cr", "credit amount"]),
        ("All OP Data Column",  st.session_state.col_op_data,  ["all op data", "op data", "party"]),
        ("Name Column",         st.session_state.col_name,     ["name", "driver name", "employee", "driver"]),
        ("Expense Type Column", st.session_state.col_exp_type, ["expense type", "type", "category", "expense"]),
    ]

    selections = {}
    c1, c2 = st.columns(2)
    for i, (label, default, hints) in enumerate(mapping_fields):
        best = _best_match(hints, default, file_cols)
        idx  = options.index(best) if best in options else 0
        with (c1 if i % 2 == 0 else c2):
            sel = st.selectbox(label, options, index=idx, key=f"map_{label}")
            selections[label] = sel

    if st.button("🚀 Load with this mapping", type="primary"):
        _apply_column_mapping(df, selections)
        st.rerun()


def _load_full_file_from_bytes(header_row: int):
    """
    FIX: load from stored bytes rather than the UploadedFile widget object,
    which becomes unavailable / unseekable on subsequent reruns.
    """
    raw_bytes = st.session_state.get("_raw_bytes")
    filename  = st.session_state.get("_raw_filename", "")
    if raw_bytes is None:
        st.error("File bytes not found in session. Please re-upload.")
        return
    try:
        buf = io.BytesIO(raw_bytes)
        if filename.lower().endswith((".xlsx", ".xls")):
            df = pd.read_excel(buf, header=header_row, dtype=str)
        else:
            df = pd.read_csv(buf, header=header_row, dtype=str, encoding="utf-8-sig")
        st.session_state["_full_raw_df"] = df
    except Exception as e:
        st.error(f"Failed to load full file: {e}")


def _apply_column_mapping(df: pd.DataFrame, selections: dict):
    """
    FIX: store column mappings in session state instead of mutating module-level
    globals (which don't survive Streamlit reruns reliably).
    """
    def _get(key):
        v = selections.get(key, "(None — skip)")
        return v if v != "(None — skip)" else None

    st.session_state.col_date     = _get("Date Column")          or st.session_state.col_date
    st.session_state.col_debit    = _get("Debit Column")         or st.session_state.col_debit
    st.session_state.col_credit   = _get("Credit Column")        or st.session_state.col_credit
    st.session_state.col_op_data  = _get("All OP Data Column")   or st.session_state.col_op_data
    st.session_state.col_name     = _get("Name Column")          or st.session_state.col_name
    st.session_state.col_exp_type = _get("Expense Type Column")  or st.session_state.col_exp_type

    # Also update module-level globals so the engine functions read the right cols
    global COL_DATE, COL_DEBIT, COL_CREDIT, COL_OP_DATA, COL_EXP_TYPE, COL_NAME
    COL_DATE     = st.session_state.col_date
    COL_DEBIT    = st.session_state.col_debit
    COL_CREDIT   = st.session_state.col_credit
    COL_OP_DATA  = st.session_state.col_op_data
    COL_EXP_TYPE = st.session_state.col_exp_type
    COL_NAME     = st.session_state.col_name

    df = df.copy()
    for col in [COL_DEBIT, COL_CREDIT]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    numeric_cols = {COL_DEBIT, COL_CREDIT}
    for col in list(df.columns):
        if col in numeric_cols:
            continue
        try:
            nan_mask     = df[col].isna()
            original_str = df[col].fillna("").astype(str)
            cleaned_str  = original_str.str.strip().str.lower()
            df[col]      = cleaned_str.where(~nan_mask, other=np.nan)
        except Exception:
            pass

    st.session_state["df_original"]      = df
    st.session_state["df_display"]       = df.copy()
    st.session_state["mapping_confirmed"] = True
    st.session_state["page"]             = 0
    st.session_state["status_msg"]       = f"✅ File loaded — {len(df):,} rows"


# ─────────────────────────────────────────────────────────
#  STATS RIBBON
# ─────────────────────────────────────────────────────────

def _render_stats_ribbon():
    df = st.session_state.df_display
    if df is None:
        return

    total = len(df)
    if "_status" in df.columns:
        exact_m   = int((df["_status"] == "✅ Matched").sum())
        consec_m  = int((df["_status"] == "📊 Consecutive Match").sum())
        sum_m     = int((df["_status"] == "🧮 Sum Match").sum())
        cross_m   = int((df["_status"] == "🔄 Cross Matched").sum())
        unmatched = int((df["_status"] == "❌ Unmatched").sum())
        matched   = exact_m + consec_m + sum_m
    else:
        matched = consec_m = sum_m = cross_m = unmatched = "—"

    cols_ss = _get_cols()
    debit_sum  = (pd.to_numeric(df[cols_ss["debit"]],  errors="coerce").sum()
                  if cols_ss["debit"]  in df.columns else 0)
    credit_sum = (pd.to_numeric(df[cols_ss["credit"]], errors="coerce").sum()
                  if cols_ss["credit"] in df.columns else 0)

    def fmt(v):
        return f"{v:,}" if isinstance(v, int) else str(v)

    cards = [
        ("ROWS",         fmt(total),     "#475569"),
        ("MATCHED",      fmt(matched),   "#16a34a"),
        ("CONSECUTIVE",  fmt(consec_m),  "#2563eb"),
        ("SUM MATCH",    fmt(sum_m),     "#7c3aed"),
        ("CROSS",        fmt(cross_m),   "#d97706"),
        ("UNMATCHED",    fmt(unmatched), "#dc2626"),
        ("TOTAL DEBIT",  f"₹{debit_sum:,.0f}",  "#0891b2"),
        ("TOTAL CREDIT", f"₹{credit_sum:,.0f}", "#3b82f6"),
    ]

    cols = st.columns(len(cards))
    for col, (label, value, color) in zip(cols, cards):
        col.markdown(
            f'<div class="stat-card">'
            f'<div class="stat-label">{label}</div>'
            f'<div class="stat-value" style="color:{color};">{value}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    st.markdown("<br>", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────
#  FILTERS
# ─────────────────────────────────────────────────────────

def _render_filters():
    df_orig = st.session_state.df_original
    if df_orig is None:
        return

    all_cols = list(df_orig.columns)

    with st.expander("🔍 Filters", expanded=True):
        fc1, fc2, fc3 = st.columns(3)

        with fc1:
            st.markdown("**Filter 1**")
            f1_col   = st.selectbox("Column##f1",  ["(none)"] + all_cols, key="f1_col")
            f1_val   = st.text_input("Value##f1",  key="f1_val")
            f1_blank = st.checkbox("Blanks only",  key="f1_blank")

        with fc2:
            st.markdown("**Filter 2**")
            f2_col   = st.selectbox("Column##f2",  ["(none)"] + all_cols, key="f2_col")
            f2_val   = st.text_input("Value##f2",  key="f2_val")
            f2_blank = st.checkbox("Blanks only",  key="f2_blank")

        with fc3:
            st.markdown("**Filter 3**")
            f3_col   = st.selectbox("Column##f3",  ["(none)"] + all_cols, key="f3_col")
            f3_val   = st.text_input("Value##f3",  key="f3_val")
            f3_blank = st.checkbox("Blanks only",  key="f3_blank")

        st.divider()
        sc1, sc2, sc3 = st.columns([2, 2, 1])

        with sc1:
            st.markdown("**Status Filter**")
            sf_matched   = st.checkbox("✅ Matched",           value=True, key="sf_matched")
            sf_consec    = st.checkbox("📊 Consecutive Match", value=True, key="sf_consec")
            sf_sum       = st.checkbox("🧮 Sum Match",         value=True, key="sf_sum")
            sf_cross     = st.checkbox("🔄 Cross Matched",     value=True, key="sf_cross")
            sf_unmatched = st.checkbox("❌ Unmatched",         value=True, key="sf_unmatched")

        with sc2:
            st.markdown("**Exclude**")
            excl_fuel      = st.checkbox("⛽ Fuel",      key="excl_fuel")
            excl_insurance = st.checkbox("🛡 Insurance", key="excl_insurance")

            # FIX: column sort control
            st.markdown("**Sort**")
            cols_ss   = _get_cols()
            sort_cols = ["(none)"] + all_cols
            sort_col  = st.selectbox("Sort by", sort_cols, key="sort_col_sel")
            sort_asc  = st.checkbox("Ascending", value=True, key="sort_asc_sel")

        with sc3:
            st.markdown("&nbsp;", unsafe_allow_html=True)
            apply_clicked = st.button("▶ Apply", type="primary", use_container_width=True)
            clear_clicked = st.button("✕ Clear",               use_container_width=True)

    status_opts = {
        "✅ Matched":           sf_matched,
        "📊 Consecutive Match": sf_consec,
        "🧮 Sum Match":         sf_sum,
        "🔄 Cross Matched":     sf_cross,
        "❌ Unmatched":         sf_unmatched,
    }

    if apply_clicked:
        _apply_filters(
            [(f1_col, f1_val, f1_blank),
             (f2_col, f2_val, f2_blank),
             (f3_col, f3_val, f3_blank)],
            status_opts, excl_fuel, excl_insurance,
            sort_col, sort_asc,
        )

    if clear_clicked:
        st.session_state.df_display = st.session_state.df_original.copy()
        st.session_state.page = 0
        st.rerun()


def _apply_filters(filter_triples, status_opts,
                   excl_fuel, excl_insurance, sort_col, sort_asc):
    df = st.session_state.df_original.copy()
    cols_ss = _get_cols()

    # Text / blank filters
    for col, val, blank in filter_triples:
        if not col or col == "(none)" or col not in df.columns:
            continue
        if blank:
            df = df[df[col].isna() |
                    df[col].astype(str).str.strip().isin(["", "nan"])]
        elif val.strip():
            if "," in val:
                parts = [p.strip() for p in val.split(",") if p.strip()]
                mask = df[col].astype(str).str.contains(parts[0], case=False, na=False)
                for part in parts[1:]:
                    mask |= df[col].astype(str).str.contains(part, case=False, na=False)
                df = df[mask]
            else:
                df = df[df[col].astype(str).str.contains(val.strip(), case=False, na=False)]

    # Exclude fuel / insurance
    exp_col = cols_ss["exp_type"]
    if exp_col in df.columns:
        excl = []
        if excl_fuel:      excl.append("fuel")
        if excl_insurance: excl.append("insurance")
        if excl:
            df = df[~df[exp_col].astype(str).str.strip().str.lower().isin(excl)]

    # Status filter
    if "_status" in df.columns:
        selected = [s for s, checked in status_opts.items() if checked]
        if len(selected) < len(status_opts):
            df = df[df["_status"].isin(selected)] if selected else df.iloc[0:0]

    # FIX: column sort
    if sort_col and sort_col != "(none)" and sort_col in df.columns:
        df = df.sort_values(sort_col, ascending=sort_asc,
                            na_position="last").reset_index(drop=True)

    st.session_state.df_display = df
    st.session_state.page = 0
    st.rerun()


# ─────────────────────────────────────────────────────────
#  DATA TABLE  (paginated + colour-coded)
# ─────────────────────────────────────────────────────────

_PAGE_SIZE = 500

def _render_table():
    df = st.session_state.df_display
    if df is None or len(df) == 0:
        st.info("No data to display.")
        return

    hidden      = st.session_state.hidden_columns
    visible_cols = [c for c in df.columns if c not in hidden]
    df_view      = df[visible_cols]

    total_rows  = len(df_view)
    total_pages = max(1, (total_rows + _PAGE_SIZE - 1) // _PAGE_SIZE)
    page        = max(0, min(st.session_state.page, total_pages - 1))
    st.session_state.page = page

    start = page * _PAGE_SIZE
    end   = min(start + _PAGE_SIZE, total_rows)
    chunk = df_view.iloc[start:end]

    # ── Pagination bar ─────────────────────────────────
    pc = st.columns([1, 1, 4, 1, 1, 2])
    with pc[0]:
        if st.button("«", key="pg_first"):
            st.session_state.page = 0; st.rerun()
    with pc[1]:
        if st.button("‹", key="pg_prev"):
            if page > 0: st.session_state.page = page - 1; st.rerun()
    with pc[2]:
        st.markdown(f"**Page {page+1} / {total_pages}** &nbsp;·&nbsp; {total_rows:,} rows shown")
    with pc[3]:
        if st.button("›", key="pg_next"):
            if page < total_pages - 1: st.session_state.page = page + 1; st.rerun()
    with pc[4]:
        if st.button("»", key="pg_last"):
            st.session_state.page = total_pages - 1; st.rerun()
    with pc[5]:
        jump = st.number_input("Go to page", min_value=1, max_value=total_pages,
                               value=page + 1, step=1,
                               label_visibility="collapsed", key="pg_jump")
        if int(jump) - 1 != page:
            st.session_state.page = int(jump) - 1; st.rerun()

    # ── Colour-coded dataframe ─────────────────────────
    if "_status" in chunk.columns:
        def highlight_row(row):
            bg = _row_highlight(str(row.get("_status", "")))
            return [bg] * len(row)
        styled = chunk.style.apply(highlight_row, axis=1)
        st.dataframe(styled, use_container_width=True, height=520)
    else:
        st.dataframe(chunk, use_container_width=True, height=520)

    st.caption(f"Showing rows {start+1:,} – {end:,} of {total_rows:,} | "
               f"Total in dataset: {len(st.session_state.df_original):,}")


# ─────────────────────────────────────────────────────────
#  SHOW MATCH  (FIX: missing from original Streamlit port)
# ─────────────────────────────────────────────────────────

def _render_show_match_section():
    df = st.session_state.df_original
    if df is None or "_match_id" not in df.columns:
        return

    st.markdown("---")
    st.markdown("### 🔗 Show Match Group")
    st.caption("Enter a Match ID (e.g. M0001 or CM001) to see all rows in that match group.")

    c1, c2 = st.columns([2, 1])
    with c1:
        mid_input = st.text_input("Match ID", key="show_match_input",
                                  placeholder="e.g. M0001")
    with c2:
        show_clicked = st.button("🔍 Show Match", key="show_match_btn")

    if show_clicked and mid_input.strip():
        mid = mid_input.strip()
        matched_df = df[df["_match_id"] == mid]
        if matched_df.empty:
            st.warning(f"No rows found with Match ID: **{mid}**")
        else:
            cols_ss = _get_cols()
            note = ""
            if "_match_note" in matched_df.columns:
                notes = matched_df["_match_note"].dropna().unique()
                if len(notes):
                    note = str(notes[0])

            st.success(f"**Match ID:** {mid} — {len(matched_df)} rows")
            if note:
                st.info(f"📋 {note}")

            show_cols = [c for c in matched_df.columns
                         if c not in ("_matched", "_match_note")]
            sub = matched_df[show_cols]

            def highlight_match_row(row):
                try:
                    d = row.get(cols_ss["debit"], None)
                    if pd.notna(d) and float(d) != 0:
                        return ["background-color: #fee2e2"] * len(row)
                except Exception:
                    pass
                return ["background-color: #dcfce7"] * len(row)

            styled = sub.style.apply(highlight_match_row, axis=1)
            st.dataframe(styled, use_container_width=True)

            # Summary
            d_col = cols_ss["debit"]
            c_col = cols_ss["credit"]
            d_total = pd.to_numeric(matched_df[d_col], errors="coerce").sum() if d_col in matched_df.columns else 0
            c_total = pd.to_numeric(matched_df[c_col], errors="coerce").sum() if c_col in matched_df.columns else 0
            sc1, sc2 = st.columns(2)
            sc1.metric("Total Debit in group",  f"₹{d_total:,.2f}")
            sc2.metric("Total Credit in group", f"₹{c_total:,.2f}")

    # FIX: also show a quick-lookup table of all match IDs present
    if "_match_id" in df.columns:
        with st.expander("📋 All Match IDs summary"):
            matched_only = df[df["_match_id"] != ""]
            if matched_only.empty:
                st.info("Run reconciliation first to see match groups.")
            else:
                summary = (matched_only.groupby("_match_id")
                           .agg(rows=("_match_id", "count"),
                                status=("_status", "first"))
                           .reset_index()
                           .rename(columns={"_match_id": "Match ID",
                                            "rows": "Row Count",
                                            "status": "Status"}))
                st.dataframe(summary, use_container_width=True, height=300)


# ─────────────────────────────────────────────────────────
#  RECONCILIATION
# ─────────────────────────────────────────────────────────

def _run_reconciliation():
    df = st.session_state.df_original
    if df is None:
        st.error("No data loaded.")
        return

    # FIX: re-sync module-level globals from session state before running engine
    global COL_DATE, COL_DEBIT, COL_CREDIT, COL_OP_DATA, COL_EXP_TYPE, COL_NAME
    COL_DATE     = st.session_state.col_date
    COL_DEBIT    = st.session_state.col_debit
    COL_CREDIT   = st.session_state.col_credit
    COL_OP_DATA  = st.session_state.col_op_data
    COL_EXP_TYPE = st.session_state.col_exp_type
    COL_NAME     = st.session_state.col_name

    with st.spinner("⏳ Running reconciliation… this may take a while for large files."):
        try:
            result = run_reconciliation(df)
            st.session_state.df_original = result
            st.session_state.df_display  = result.copy()
            st.session_state.status_msg  = "✅ Reconciliation complete!"
            st.session_state.page        = 0
        except Exception as e:
            import traceback
            st.session_state.status_msg = f"❌ Reconciliation failed: {e}\n{traceback.format_exc()}"
    st.rerun()


# ─────────────────────────────────────────────────────────
#  EDIT ACTIONS
# ─────────────────────────────────────────────────────────

def _edit_add_row(df):
    st.markdown("**Add a new row**")
    user_cols = [c for c in df.columns if not c.startswith("_")]
    with st.form("add_row_form", clear_on_submit=True):
        vals = {}
        for col in user_cols:
            vals[col] = st.text_input(col, key=f"ar_{col}")
        position  = st.radio("Insert at", ["End", "Beginning"], horizontal=True)
        submitted = st.form_submit_button("💾 Add Row")
    if submitted:
        new_row = pd.DataFrame([vals])
        orig    = st.session_state.df_original
        if position == "Beginning":
            st.session_state.df_original = pd.concat([new_row, orig], ignore_index=True)
        else:
            st.session_state.df_original = pd.concat([orig, new_row], ignore_index=True)
        st.session_state.df_display = st.session_state.df_original.copy()
        st.session_state.status_msg = "✅ Row added."
        st.rerun()


def _edit_edit_row(df):
    """FIX: Edit a specific row by its 0-based index."""
    st.markdown("**Edit a row**")
    total = len(df)
    st.caption(f"Total rows: {total:,} (0-based index)")

    row_idx = st.number_input("Row index to edit", min_value=0,
                              max_value=max(0, total - 1),
                              value=0, step=1, key="edit_row_idx")

    user_cols = [c for c in df.columns if not c.startswith("_")]
    current   = df.iloc[int(row_idx)]

    with st.form("edit_row_form"):
        vals = {}
        for col in user_cols:
            cur_val = str(current[col]) if pd.notna(current[col]) else ""
            vals[col] = st.text_input(col, value=cur_val, key=f"er_{col}")
        submitted = st.form_submit_button("💾 Save Changes")

    if submitted:
        for col, val in vals.items():
            if col in st.session_state.df_original.columns:
                # Preserve dtype
                try:
                    col_dtype = st.session_state.df_original[col].dtype
                    if pd.api.types.is_float_dtype(col_dtype):
                        converted = float(val) if val.strip() else np.nan
                    elif pd.api.types.is_integer_dtype(col_dtype):
                        converted = int(float(val)) if val.strip() else np.nan
                    else:
                        converted = val if val.strip() else np.nan
                except (ValueError, TypeError):
                    converted = val
                st.session_state.df_original.iat[int(row_idx), df.columns.get_loc(col)] = converted
        st.session_state.df_display = st.session_state.df_original.copy()
        st.session_state.status_msg = f"✅ Row {row_idx} updated."
        st.rerun()


def _edit_delete_row(df):
    st.markdown("**Delete rows by index**")
    st.caption(f"Total rows: {len(df):,}. Comma-separated 0-based indices, or a range like 5-10.")
    with st.form("del_row_form"):
        idx_input = st.text_input("Row indices (e.g. 0, 5, 10 or 3-7)", key="del_idx")
        submitted = st.form_submit_button("🗑 Delete Rows")
    if submitted:
        try:
            idxs = set()
            for part in idx_input.split(","):
                part = part.strip()
                if "-" in part:
                    lo, hi = part.split("-", 1)
                    idxs.update(range(int(lo.strip()), int(hi.strip()) + 1))
                elif part:
                    idxs.add(int(part))
            idxs = sorted(idxs)
            st.session_state.df_original = df.drop(index=idxs).reset_index(drop=True)
            st.session_state.df_display  = st.session_state.df_original.copy()
            st.session_state.status_msg  = f"✅ Deleted {len(idxs)} row(s)."
            st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")


def _edit_bulk(df):
    st.markdown("**Bulk edit a column**")
    visible_cols = [c for c in df.columns if not c.startswith("_")]
    with st.form("bulk_edit_form"):
        col_sel   = st.selectbox("Column to edit", visible_cols, key="be_col")
        new_val   = st.text_input("New value for all selected rows", key="be_val")
        scope     = st.radio("Apply to", ["Filtered rows only", "All rows"], horizontal=True)
        submitted = st.form_submit_button("✅ Apply")
    if submitted:
        if scope == "Filtered rows only" and st.session_state.df_display is not None:
            idxs = st.session_state.df_display.index.tolist()
            st.session_state.df_original.loc[idxs, col_sel] = new_val
        else:
            st.session_state.df_original[col_sel] = new_val
        st.session_state.df_display = st.session_state.df_original.copy()
        st.session_state.status_msg = f"✅ '{col_sel}' updated."
        st.rerun()


def _edit_add_column(df):
    st.markdown("**Add computed column**")
    st.caption("Use `df` in your code. Assign result to `result` for multi-line blocks.")
    with st.form("add_col_form"):
        col_name  = st.text_input("New column name", key="ac_name")
        code      = st.text_area("Python expression / code", height=120, key="ac_code",
                                 placeholder="e.g.  df['Debit'].fillna(0) * 1.1")
        submitted = st.form_submit_button("✅ Apply")
    if submitted:
        if not col_name.strip():
            st.error("Please enter a column name.")
            return
        try:
            ns = {
                "df": df.copy(), "pd": pd, "np": np, "result": None,
                "str": str, "int": int, "float": float, "len": len,
                "round": round, "abs": abs, "min": min, "max": max,
                "list": list, "dict": dict, "range": range,
                "True": True, "False": False, "None": None,
            }
            try:
                result = eval(code.strip(), ns, ns)
            except SyntaxError:
                exec(code.strip(), ns, ns)
                result = ns.get("result")
                if result is None:
                    raise ValueError("Multi-line code must assign its output to 'result'.")
            if not hasattr(result, "__len__") or isinstance(result, str):
                result = pd.Series([result] * len(df))
            st.session_state.df_original[col_name] = result.values
            st.session_state.df_display = st.session_state.df_original.copy()
            st.session_state.status_msg = f"✅ Column '{col_name}' added."
            st.rerun()
        except Exception as e:
            st.error(f"Error computing column: {e}")


def _edit_hide_show(df):
    st.markdown("**Hide / show columns**")
    all_cols = list(df.columns)
    hidden   = st.session_state.hidden_columns
    to_hide  = st.multiselect(
        "Columns to hide (data is preserved)",
        options=all_cols,
        default=sorted(hidden),
        key="hide_cols_ms",
    )
    if st.button("Apply visibility", key="apply_vis"):
        st.session_state.hidden_columns = set(to_hide)
        st.rerun()


def _edit_delete_columns(df):
    st.markdown("**Permanently delete columns**")
    st.warning("⚠️ This cannot be undone.")
    to_del = st.multiselect("Columns to delete permanently",
                             options=list(df.columns),
                             key="del_cols_ms")
    if st.button("🗑 Delete Selected", type="primary", key="do_del_cols"):
        if to_del:
            st.session_state.df_original.drop(columns=to_del, inplace=True, errors="ignore")
            st.session_state.df_display = st.session_state.df_original.copy()
            st.session_state.hidden_columns.difference_update(set(to_del))
            st.session_state.status_msg = f"✅ Deleted {len(to_del)} column(s)."
            st.rerun()


# ══════════════════════════════════════════════════════════
if __name__ == "__main__":
    main()
