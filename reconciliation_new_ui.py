"""
Account Reconciliation Tool
- Load large Excel/CSV files (5L+ rows)
- View, Edit, Add, Delete rows
- Triple simultaneous column filter
- Alternating Reconciliation:
    * Sort by date
    * Check first entry — Debit or Credit
    * If Debit first  → find next Credit(s) whose SUM == that Debit amount
    * If Credit first → find next Debit(s)  whose SUM == that Credit amount
    * After each match, flip side and repeat from next unmatched entry
    * Number of entries in a combination is unknown / variable
- Reconcile at All OP Data + Expense Type level separately
"""

import streamlit as st
import io
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import time
from itertools import combinations

# ─────────────────────────────────────────────
# CONFIG — column name mappings (edit if needed)
# ─────────────────────────────────────────────
COL_DATE      = "Post Date"
COL_DEBIT     = "Debit"
COL_CREDIT    = "Credit"
COL_OP_DATA   = "All OP Data"
COL_EXP_TYPE  = "Expense Type"
COL_NAME      = "Name"

# Expense types that must ONLY reconcile within the same type (strict 1-to-1 category)
STRICT_EXPENSE_TYPES = {"fuel", "insurance", "camera", "prepass", "plates"}

PAGE_SIZE = 500



# ══════════════════════════════════════════════════════════
#  RECONCILIATION ENGINE  (optimised — array-based, int math)
#  Phase 1:   Exact 1:1 matching (debit == credit)
#  Phase 1.5: Consecutive N:1 sum matching (multi-pass)
#  Phase 1.6: Greedy accumulation (all-remaining → single target)
#  Phase 1.7: N:M consecutive matching (window vs window)
#  Phase 2A:  Meet-in-the-middle subset-sum (≤25 entries, any combo)
#  Phase 2B:  Combinatorial fallback (larger groups, combo ≤5)
# ══════════════════════════════════════════════════════════



def reconcile_group(group_df: pd.DataFrame, match_counter_start: int = 1):
    """
    Reconcile a single group using multi-phase matching.

    Phase 1:   Exact 1:1 — debit == credit (across entire group)
    Phase 1.5: Consecutive sum — contiguous sequence of entries on
               one side sums to a single entry on the opposite side

    Returns (df, next_match_counter).
    """
    GROUP_TIME_LIMIT = 30.0  # hard cap: 30 seconds per group
    _group_start = time.perf_counter()

    def _group_timed_out():
        return (time.perf_counter() - _group_start) > GROUP_TIME_LIMIT

    df = group_df.copy().reset_index(drop=True)
    df[COL_DATE] = pd.to_datetime(df[COL_DATE], errors="coerce")
    df = df.sort_values([COL_DATE], kind="mergesort").reset_index(drop=True)

    n = len(df)

    # ── Vectorised pre-extraction ──
    debit_arr  = [0] * n
    credit_arr = [0] * n

    if COL_DEBIT in df.columns:
        dvals = pd.to_numeric(df[COL_DEBIT], errors="coerce").fillna(0.0).values
        for i in range(n):
            if dvals[i] != 0.0:
                debit_arr[i] = int(round(dvals[i] * 100))

    if COL_CREDIT in df.columns:
        cvals = pd.to_numeric(df[COL_CREDIT], errors="coerce").fillna(0.0).values
        for i in range(n):
            if cvals[i] != 0.0:
                credit_arr[i] = int(round(cvals[i] * 100))

    # Result arrays
    matched    = [False] * n
    match_id   = [""] * n
    match_note = [""] * n
    match_type = [""] * n      # "exact" for 1:1, "consecutive" for consecutive-sum
    used       = [False] * n
    match_counter = match_counter_start

    # ──────────────────────────────────────────────────────────────────
    # PHASE 1: Exact 1:1 matches (debit value == credit value)
    # Build a map of credit values → list of indices for O(1) lookup
    # ──────────────────────────────────────────────────────────────────
    credit_map = {}  # paise_value → [list of indices]
    for i in range(n):
        if credit_arr[i] != 0:
            credit_map.setdefault(credit_arr[i], []).append(i)

    for i in range(n):
        if used[i] or debit_arr[i] == 0:
            continue
        d_val = debit_arr[i]
        if d_val in credit_map:
            # Find first unused credit with this value
            clist = credit_map[d_val]
            for ci in clist:
                if not used[ci]:
                    # 1:1 match found
                    mid = f"M{match_counter:04d}"
                    match_counter += 1
                    amount = d_val / 100.0
                    note = f"Debit ₹{amount:,.2f} ↔ Credit ₹{amount:,.2f} (1:1)"

                    for idx in [i, ci]:
                        matched[idx]    = True
                        match_id[idx]   = mid
                        match_note[idx] = note
                        match_type[idx] = "exact"
                        used[idx]       = True
                    break

    # ──────────────────────────────────────────────────────────────────
    # PHASE 1.5: CONSECUTIVE SUM MATCHING  (multi-pass until stable)
    #   Find contiguous/consecutive sequences of debit entries whose
    #   sum equals a single credit entry, and vice versa.
    #   Runs repeatedly: after each round of matches, newly-adjacent
    #   entries may form new consecutive windows.
    # ──────────────────────────────────────────────────────────────────

    def _find_consecutive_sum_matches(source_arr, target_arr, source_label, target_label):
        """Find consecutive sequences in source_arr that sum to a target in target_arr.
        Returns number of new matches found (0 means done).
        """
        nonlocal match_counter
        matches_found = 0

        # Collect unused source indices (in order)
        src_indices = [i for i in range(n) if source_arr[i] != 0 and not used[i]]
        if not src_indices:
            return 0

        # Collect unused target values → indices for lookup
        tgt_map = {}  # paise_value → [list of indices]
        for i in range(n):
            if target_arr[i] != 0 and not used[i]:
                tgt_map.setdefault(target_arr[i], []).append(i)
        if not tgt_map:
            return 0

        for l in range(len(src_indices)):
            if used[src_indices[l]]:
                continue
            if _group_timed_out():
                return matches_found
            running_sum = 0
            for r in range(l, len(src_indices)):
                if r > l:
                    gap_has_source = False
                    for g in range(src_indices[r-1] + 1, src_indices[r]):
                        if source_arr[g] != 0 and not used[g]:
                            gap_has_source = True
                            break
                    if gap_has_source:
                        break

                idx_r = src_indices[r]
                if used[idx_r]:
                    break
                running_sum += source_arr[idx_r]
                count = r - l + 1

                if count < 2:
                    continue

                if running_sum in tgt_map:
                    tgt_list = tgt_map[running_sum]
                    matched_tgt = None
                    for ti in tgt_list:
                        if not used[ti]:
                            matched_tgt = ti
                            break
                    if matched_tgt is not None:
                        mid = f"M{match_counter:04d}"
                        match_counter += 1
                        amount = running_sum / 100.0
                        consec_indices = [src_indices[k] for k in range(l, r + 1)]
                        note = (f"{count} consecutive {source_label} entries "
                                f"₹{amount:,.2f} ↔ {target_label} ₹{amount:,.2f} "
                                f"(consecutive sum)")

                        all_indices = consec_indices + [matched_tgt]
                        for idx in all_indices:
                            matched[idx]    = True
                            match_id[idx]   = mid
                            match_note[idx] = note
                            match_type[idx] = "consecutive"
                            used[idx]       = True

                        tgt_list.remove(matched_tgt)
                        if not tgt_list:
                            del tgt_map[running_sum]

                        matches_found += 1
                        break
        return matches_found

    # Run consecutive N:1 matching in multiple passes until stable
    for _pass in range(20):
        if _group_timed_out():
            break
        found = 0
        found += _find_consecutive_sum_matches(debit_arr, credit_arr, "debit", "Credit")
        found += _find_consecutive_sum_matches(credit_arr, debit_arr, "credit", "Debit")
        if found == 0:
            break

    # ──────────────────────────────────────────────────────────────────
    # PHASE 1.6: GREEDY ACCUMULATION  (all-remaining → single target)
    #   For each unmatched entry on one side, check if ALL remaining
    #   unmatched entries on the other side sum to it.  Catches large
    #   N:1 patterns (e.g. 14 debits → 1 credit) in O(n).
    # ──────────────────────────────────────────────────────────────────

    def _greedy_all_remaining(source_arr, target_arr, source_label, target_label):
        """Check if all remaining unused source entries sum to a single target."""
        nonlocal match_counter
        matches_found = 0

        src_indices = [i for i in range(n) if source_arr[i] != 0 and not used[i]]
        if len(src_indices) < 2:
            return 0
        src_total = sum(source_arr[i] for i in src_indices)

        for i in range(n):
            if target_arr[i] != 0 and not used[i] and target_arr[i] == src_total:
                mid = f"M{match_counter:04d}"
                match_counter += 1
                amount = src_total / 100.0
                note = (f"{len(src_indices)} {source_label} entries "
                        f"₹{amount:,.2f} ↔ {target_label} ₹{amount:,.2f} "
                        f"(sum match)")
                all_matched = src_indices + [i]
                for idx in all_matched:
                    matched[idx]    = True
                    match_id[idx]   = mid
                    match_note[idx] = note
                    match_type[idx] = "sum"
                    used[idx]       = True
                matches_found += 1
                break  # all source entries are now used
        return matches_found

    for _pass in range(10):
        if _group_timed_out():
            break
        found = 0
        found += _greedy_all_remaining(debit_arr, credit_arr, "debit", "Credit")
        found += _greedy_all_remaining(credit_arr, debit_arr, "credit", "Debit")
        if found == 0:
            break

    # ──────────────────────────────────────────────────────────────────
    # PHASE 1.7: N:M CONSECUTIVE MATCHING
    #   Contiguous window of debits whose sum == contiguous window of
    #   credits.  Both windows must have ≥ 2 entries (N:1 already done).
    # ──────────────────────────────────────────────────────────────────

    MAX_WINDOWS = 50_000  # cap to prevent memory/time blowup

    def _get_consecutive_windows(arr, min_size=1, max_size=15):
        """Return list of (indices_list, window_sum) for consecutive unused entries."""
        windows = []
        idxs = [i for i in range(n) if arr[i] != 0 and not used[i]]
        if len(idxs) < min_size:
            return windows
        for l in range(len(idxs)):
            if _group_timed_out() or len(windows) >= MAX_WINDOWS:
                break
            if used[idxs[l]]:
                continue
            running = 0
            for r in range(l, min(l + max_size, len(idxs))):
                if r > l:
                    gap = False
                    for g in range(idxs[r-1] + 1, idxs[r]):
                        if arr[g] != 0 and not used[g]:
                            gap = True
                            break
                    if gap:
                        break
                if used[idxs[r]]:
                    break
                running += arr[idxs[r]]
                sz = r - l + 1
                if sz >= min_size:
                    windows.append(([idxs[k] for k in range(l, r + 1)], running))
        return windows

    # Build debit and credit windows, match by sum
    if not _group_timed_out():
        d_windows = _get_consecutive_windows(debit_arr, min_size=2, max_size=10)
        c_win_map = {}  # sum → list of index-lists
        for c_idx_list, c_sum in _get_consecutive_windows(credit_arr, min_size=2, max_size=10):
            c_win_map.setdefault(c_sum, []).append(c_idx_list)

        for d_idx_list, d_sum in d_windows:
            if _group_timed_out():
                break
            if any(used[i] for i in d_idx_list):
                continue
            if d_sum not in c_win_map:
                continue
            for c_idx_list in c_win_map[d_sum]:
                if any(used[i] for i in c_idx_list):
                    continue
                # N:M match found
                mid = f"M{match_counter:04d}"
                match_counter += 1
                amount = d_sum / 100.0
                note = (f"{len(d_idx_list)} debit + {len(c_idx_list)} credit entries "
                        f"₹{amount:,.2f} (N:M consecutive)")
                for idx in d_idx_list + c_idx_list:
                    matched[idx]    = True
                    match_id[idx]   = mid
                    match_note[idx] = note
                    match_type[idx] = "consecutive"
                    used[idx]       = True
                break

    # ──────────────────────────────────────────────────────────────────
    # PHASE 2: GLOBAL (NON-CONSECUTIVE) SUM MATCHING
    #   Strategy A — Meet-in-the-middle subset-sum for ≤ MITM_MAX
    #     source entries: O(2^(n/2)), finds ANY subset that sums to
    #     a target.  Handles 6–25 entry combos that combinatorial
    #     approach cannot reach.
    #   Strategy B — Combinatorial fallback for larger groups.
    # ──────────────────────────────────────────────────────────────────

    MITM_MAX = 25              # max source entries for meet-in-the-middle
    MAX_COMBO = 5              # fallback combo size for large groups
    MAX_SRC_BIG = 80
    MAX_ITERATIONS = 500_000
    TIME_LIMIT = 10.0

    _phase2_start = time.perf_counter()  # separate timer for Phase 2

    # ── Strategy A: Meet-in-the-middle subset-sum ─────────────────────

    MITM_TIME_LIMIT = 5.0   # seconds per MITM call

    def _mitm_find_subset(entries, target):
        """Find a subset (size ≥ 2) of entries that sums exactly to target.
        entries: list of (index, value).  Returns list of indices or None.
        Uses meet-in-the-middle: O(2^(n/2)) time and space.
        """
        ne = len(entries)
        if ne < 2:
            return None
        _mitm_start = time.perf_counter()
        mid_pt = ne // 2
        left, right = entries[:mid_pt], entries[mid_pt:]

        # All non-empty subsets of left half → {sum: (indices, count)}
        left_sums = {}  # sum → (indices_list, subset_size)
        for mask in range(1, 1 << len(left)):
            if mask % 10_000 == 0 and (time.perf_counter() - _mitm_start) > MITM_TIME_LIMIT:
                return None  # bail out
            s, idxs = 0, []
            for i in range(len(left)):
                if mask & (1 << i):
                    s += left[i][1]
                    idxs.append(left[i][0])
            if s <= target and s not in left_sums:
                left_sums[s] = idxs

        # Left-only hit
        if target in left_sums and len(left_sums[target]) >= 2:
            return left_sums[target]

        # Right subsets, check complement in left
        for mask in range(1, 1 << len(right)):
            if mask % 10_000 == 0 and (time.perf_counter() - _mitm_start) > MITM_TIME_LIMIT:
                return None  # bail out
            s, idxs = 0, []
            for i in range(len(right)):
                if mask & (1 << i):
                    s += right[i][1]
                    idxs.append(right[i][0])
            if s > target:
                continue
            if s == target and len(idxs) >= 2:
                return idxs
            complement = target - s
            if complement in left_sums:
                combined = left_sums[complement] + idxs
                if len(combined) >= 2:
                    return combined
        return None

    def _find_global_mitm(source_arr, target_arr, source_label, target_label):
        """Meet-in-the-middle global sum matching for small groups."""
        nonlocal match_counter
        src = [(i, source_arr[i]) for i in range(n)
               if source_arr[i] != 0 and not used[i]]
        if len(src) < 2 or len(src) > MITM_MAX:
            return False

        targets = [(i, target_arr[i]) for i in range(n)
                   if target_arr[i] != 0 and not used[i]]
        if not targets:
            return False

        found_any = False
        for tgt_idx, tgt_val in targets:
            if used[tgt_idx]:
                continue
            src = [(i, v) for i, v in src if not used[i]]
            if len(src) < 2:
                break
            result = _mitm_find_subset(src, tgt_val)
            if result is not None:
                mid = f"M{match_counter:04d}"
                match_counter += 1
                amount = tgt_val / 100.0
                note = (f"{len(result)} {source_label} entries "
                        f"₹{amount:,.2f} ↔ {target_label} ₹{amount:,.2f} "
                        f"(sum match)")
                for idx in result + [tgt_idx]:
                    matched[idx]    = True
                    match_id[idx]   = mid
                    match_note[idx] = note
                    match_type[idx] = "sum"
                    used[idx]       = True
                found_any = True
        return found_any

    # Run MITM in passes until stable
    for _mitm_pass in range(10):
        if _group_timed_out():
            break
        found_mitm = False
        d_rem = sum(1 for i in range(n) if debit_arr[i] != 0 and not used[i])
        c_rem = sum(1 for i in range(n) if credit_arr[i] != 0 and not used[i])
        if 2 <= d_rem <= MITM_MAX:
            found_mitm |= _find_global_mitm(debit_arr, credit_arr, "debit", "Credit")
        if 2 <= c_rem <= MITM_MAX:
            found_mitm |= _find_global_mitm(credit_arr, debit_arr, "credit", "Debit")
        if not found_mitm:
            break

    # ── Strategy B: Combinatorial fallback for remaining/larger groups ─

    def _find_global_sum_matches(source_arr, target_arr, source_label, target_label):
        nonlocal match_counter
        if _group_timed_out():
            return
        _phase2b_start = time.perf_counter()  # own timer for this call
        src = [(i, source_arr[i]) for i in range(n) if source_arr[i] != 0 and not used[i]]
        if len(src) < 2:
            return

        tgt_map = {}
        for i in range(n):
            if target_arr[i] != 0 and not used[i]:
                tgt_map.setdefault(target_arr[i], []).append(i)
        if not tgt_map:
            return

        effective_max = MAX_COMBO
        if len(src) > MAX_SRC_BIG:
            effective_max = min(effective_max, 3)

        iteration_count = 0

        for combo_size in range(2, effective_max + 1):
            src = [(i, v) for i, v in src if not used[i]]
            if len(src) < combo_size:
                break

            for combo in combinations(src, combo_size):
                iteration_count += 1
                if iteration_count > MAX_ITERATIONS:
                    return
                if iteration_count % 10_000 == 0:
                    if time.perf_counter() - _phase2b_start > TIME_LIMIT:
                        return
                    if _group_timed_out():
                        return

                if any(used[idx] for idx, _ in combo):
                    continue
                total = sum(v for _, v in combo)
                if total not in tgt_map:
                    continue
                tgt_list = tgt_map[total]
                matched_tgt = None
                for ti in tgt_list:
                    if not used[ti]:
                        matched_tgt = ti
                        break
                if matched_tgt is None:
                    continue

                mid = f"M{match_counter:04d}"
                match_counter += 1
                amount = total / 100.0
                combo_indices = [idx for idx, _ in combo]
                note = (f"{combo_size} {source_label} entries "
                        f"₹{amount:,.2f} ↔ {target_label} ₹{amount:,.2f} "
                        f"(sum match)")

                for idx in combo_indices + [matched_tgt]:
                    matched[idx]    = True
                    match_id[idx]   = mid
                    match_note[idx] = note
                    match_type[idx] = "sum"
                    used[idx]       = True

                tgt_list.remove(matched_tgt)
                if not tgt_list:
                    del tgt_map[total]

    if not _group_timed_out():
        _find_global_sum_matches(debit_arr, credit_arr, "debit", "Credit")
    if not _group_timed_out():
        _find_global_sum_matches(credit_arr, debit_arr, "credit", "Debit")

    # Write results back in bulk
    df["_matched"]    = matched
    df["_match_id"]   = match_id
    df["_match_note"] = match_note
    df["_match_type"] = match_type

    # Assign status
    statuses = []
    for i in range(n):
        if matched[i] and match_type[i] == "sum":
            statuses.append("🧮 Sum Match")
        elif matched[i] and match_type[i] == "consecutive":
            statuses.append("📊 Consecutive Match")
        elif matched[i]:
            statuses.append("✅ Matched")
        else:
            statuses.append("❌ Unmatched")
    df["_status"] = statuses
    return df, match_counter


def _row_group_key(row, has_op_data: bool, has_name: bool, has_exp_type: bool) -> tuple:
    """
    Build a (entity_key, expense_group) tuple for a row.

    Entity key priority:
      1. All OP Data  (if non-blank)
      2. Name         (fallback when All OP Data is blank/null)
      3. '__unknown__'

    Expense group:
      - Strict types (fuel, insurance, camera, prepass, plates) → exact type name
      - All other types → '__other__'  (can cross-match freely)
    """
    # ── Entity ──────────────────────────────────────────────────────────────
    entity = "__unknown__"
    if has_op_data:
        op = row[COL_OP_DATA]
        if pd.notna(op) and str(op).strip() not in ("", "nan"):
            entity = str(op).strip().lower()
    if entity == "__unknown__" and has_name:
        nm = row[COL_NAME]
        if pd.notna(nm) and str(nm).strip() not in ("", "nan"):
            entity = str(nm).strip().lower()

    # ── Expense group ────────────────────────────────────────────────────────
    exp_group = "__other__"
    if has_exp_type:
        raw = row[COL_EXP_TYPE]
        exp = str(raw).strip().lower() if pd.notna(raw) and str(raw).strip() not in ("", "nan") else ""
        if exp in STRICT_EXPENSE_TYPES:
            exp_group = exp          # strict: keep exact type so they never cross-match
        # else: leave as '__other__' — all non-strict types pool together

    return (entity, exp_group)


def run_reconciliation(df: pd.DataFrame, progress_cb=None) -> pd.DataFrame:
    """
    New grouping logic:
      • Entity = All OP Data  (primary)  →  Name  (fallback if OP Data is blank)
      • Expense group:
          - fuel / insurance / camera / prepass / plates  →  strict (same-type only)
          - everything else (advance, repair, eps, paycheck, void, …)  → '__other__'
            these pool together and can cross-match freely within the same entity

    progress_cb: optional callable(msg: str) for UI progress updates.
    """
    has_op_data  = COL_OP_DATA  in df.columns
    has_name     = COL_NAME     in df.columns
    has_exp_type = COL_EXP_TYPE in df.columns

    if not has_op_data and not has_name:
        if progress_cb:
            progress_cb("Reconciling (single group)…")
        result, _ = reconcile_group(df)
        return result

    df = df.copy()

    # ── Vectorised group-key construction (replaces slow df.apply) ────────
    # Entity: primary = All OP Data, fallback = Name
    entity = pd.Series("__unknown__", index=df.index)
    if has_op_data:
        op = df[COL_OP_DATA].fillna("").astype(str).str.strip().str.lower()
        mask_op = ~op.isin(["", "nan"])
        entity = entity.where(~mask_op, op)
    if has_name:
        nm = df[COL_NAME].fillna("").astype(str).str.strip().str.lower()
        mask_nm = (entity == "__unknown__") & ~nm.isin(["", "nan"])
        entity = entity.where(~mask_nm, nm)

    # Expense group: strict types stay as-is, everything else → '__other__'
    exp_group = pd.Series("__other__", index=df.index)
    if has_exp_type:
        exp_raw = df[COL_EXP_TYPE].fillna("").astype(str).str.strip().str.lower()
        is_strict = exp_raw.isin(STRICT_EXPENSE_TYPES)
        exp_group = exp_group.where(~is_strict, exp_raw)

    df["_grp"] = list(zip(entity, exp_group))

    groups = list(df.groupby("_grp", dropna=False, sort=False))
    total_groups = len(groups)
    results = []
    global_counter = 1
    failed_groups = 0
    for idx, (_key, grp) in enumerate(groups, 1):
        if progress_cb and (idx % 5 == 0 or idx == 1):
            progress_cb(f"Processing group {idx}/{total_groups} ({len(grp)} rows)…")
        try:
            reconciled, global_counter = reconcile_group(
                grp.drop(columns=["_grp"]), match_counter_start=global_counter
            )
            results.append(reconciled)
        except Exception:
            # If a single group fails, keep its rows unmatched rather than
            # crashing the entire reconciliation.
            failed_grp = grp.drop(columns=["_grp"]).copy()
            failed_grp["_matched"]    = False
            failed_grp["_match_id"]   = ""
            failed_grp["_match_note"] = ""
            failed_grp["_match_type"] = ""
            failed_grp["_status"]     = "❌ Unmatched"
            results.append(failed_grp)
            failed_groups += 1

    result = pd.concat(results, ignore_index=True)

    if progress_cb and failed_groups > 0:
        progress_cb(f"⚠ {failed_groups} group(s) could not be reconciled — marked Unmatched")

    # ── POST-PROCESS: relabel cross-type matches as CM001, CM002, … ──────
    if has_exp_type and COL_EXP_TYPE in result.columns:
        if progress_cb:
            progress_cb("Detecting cross-type matches…")
        cm_counter = 1
        # Vectorised: group by match_id quickly
        matched_mask = result["_match_id"] != ""
        if matched_mask.any():
            id_groups = result.loc[matched_mask].groupby("_match_id", sort=False)
            for mid, grp_rows in id_groups:
                exp_types = set(
                    str(t).strip().lower()
                    for t in grp_rows[COL_EXP_TYPE].dropna()
                    if str(t).strip().lower() not in ("", "nan")
                )
                if len(exp_types) > 1:
                    mask = result["_match_id"] == mid
                    new_id = f"CM{cm_counter:03d}"
                    cm_counter += 1
                    cross_types = ", ".join(sorted(exp_types))
                    old_note = grp_rows["_match_note"].iloc[0] if "_match_note" in grp_rows.columns else ""
                    new_note = f"{old_note} [Cross: {cross_types}]"
                    result.loc[mask, "_match_id"]   = new_id
                    result.loc[mask, "_match_note"] = new_note
                    result.loc[mask, "_status"]     = "🔄 Cross Matched"

    return result


# ══════════════════════════════════════════════════════════
#  STREAMLIT APPLICATION  —  AccountSync
# ══════════════════════════════════════════════════════════

st.set_page_config(
    page_title="AccountSync · Reconciliation",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ────────────────────────────────────────────
st.markdown("""
<style>
    .block-container { padding-top: 1rem; padding-bottom: 0; }
    div[data-testid="stMetric"] {
        background: #f7f8fa; border: 1px solid #e2e8f0;
        border-radius: 8px; padding: 8px 14px;
        box-shadow: 0 1px 2px rgba(0,0,0,0.04);
    }
    div[data-testid="stMetric"] label { font-size: 0.65rem !important; color: #64748b; }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] { font-size: 1.15rem !important; }
    section[data-testid="stSidebar"] { background: #f7f8fa; }
    .stDataFrame { border-radius: 8px; }
</style>
""", unsafe_allow_html=True)


# ── Session State Initialisation ──────────────────────────
def _init_session():
    defaults = {
        "df_original": None,
        "df_display": None,
        "current_page": 0,
        "file_name": "",
        "step": "upload",          # upload | header_select | column_map | ready
        "raw_df": None,
        "preview_df": None,
        "selected_header": 0,
        "hidden_columns": set(),
        "_uploaded_bytes": None,
        "_last_upload": "",
        # Column mapping (persisted so engine functions can use them)
        "_col_date": COL_DATE,
        "_col_debit": COL_DEBIT,
        "_col_credit": COL_CREDIT,
        "_col_op_data": COL_OP_DATA,
        "_col_name": COL_NAME,
        "_col_exp_type": COL_EXP_TYPE,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

_init_session()


# ── Helpers ───────────────────────────────────────────────

def _best_match(file_cols, hints, current_default):
    """Find the best matching column name from a list of hints."""
    if current_default in file_cols:
        return current_default
    lower_map = {c.lower().strip(): c for c in file_cols}
    for h in hints:
        if h.lower() in lower_map:
            return lower_map[h.lower()]
    for h in hints:
        for lc, orig in lower_map.items():
            if h.lower() in lc or lc in h.lower():
                return orig
    return file_cols[0] if file_cols else ""


def _apply_filters():
    """Apply triple-column, status, and exclude filters to df_original."""
    global COL_DATE, COL_DEBIT, COL_CREDIT, COL_OP_DATA, COL_EXP_TYPE, COL_NAME
    COL_DATE = st.session_state._col_date
    COL_DEBIT = st.session_state._col_debit
    COL_CREDIT = st.session_state._col_credit
    COL_OP_DATA = st.session_state._col_op_data
    COL_NAME = st.session_state._col_name
    COL_EXP_TYPE = st.session_state._col_exp_type

    if st.session_state.df_original is None:
        return
    df = st.session_state.df_original.copy()

    for i in range(3):
        col = st.session_state.get(f"fcol_{i}", "")
        val = st.session_state.get(f"fval_{i}", "").strip()
        blank = st.session_state.get(f"fblank_{i}", False)
        if not col or col not in df.columns:
            continue
        if blank:
            df = df[
                df[col].isna()
                | (df[col].astype(str).str.strip() == "")
                | (df[col].astype(str).str.strip() == "nan")
            ]
        elif val:
            if "," in val:
                parts = [p.strip() for p in val.split(",") if p.strip()]
                if parts:
                    mask = df[col].astype(str).str.contains(parts[0], case=False, na=False)
                    for p in parts[1:]:
                        mask = mask | df[col].astype(str).str.contains(p, case=False, na=False)
                    df = df[mask]
            else:
                df = df[df[col].astype(str).str.contains(val, case=False, na=False)]

    if COL_EXP_TYPE in df.columns:
        excludes = []
        if st.session_state.get("exclude_fuel", False):
            excludes.append("fuel")
        if st.session_state.get("exclude_insurance", False):
            excludes.append("insurance")
        if excludes:
            df = df[~df[COL_EXP_TYPE].astype(str).str.strip().str.lower().isin(excludes)]

    if "_status" in df.columns:
        all_keys = [
            "✅ Matched", "📊 Consecutive Match",
            "🧮 Sum Match", "🔄 Cross Matched", "❌ Unmatched",
        ]
        selected = [k for k in all_keys if st.session_state.get(f"st_{k}", True)]
        if len(selected) < len(all_keys):
            if selected:
                df = df[df["_status"].isin(selected)]
            else:
                df = df.iloc[0:0]

    st.session_state.df_display = df
    st.session_state.current_page = 0


def _export_styled_excel(df_export):
    """Generate colour-coded Excel bytes for download."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Reconciliation"

    hdr_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    hdr_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
    fills = {
        "matched":  (PatternFill("solid", fgColor="D5F5E3"), Font(name="Segoe UI", color="1A6B3C", size=10)),
        "consec":   (PatternFill("solid", fgColor="D6EAF8"), Font(name="Segoe UI", color="1A4D8E", size=10)),
        "sum":      (PatternFill("solid", fgColor="E8DAEF"), Font(name="Segoe UI", color="6C3483", size=10)),
        "cross":    (PatternFill("solid", fgColor="FEF9E7"), Font(name="Segoe UI", color="7D6608", size=10)),
        "unmatched":(PatternFill("solid", fgColor="FADBD8"), Font(name="Segoe UI", color="922B21", size=10)),
    }
    default_font = Font(name="Segoe UI", size=10)
    border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

    cols = list(df_export.columns)
    si = next((i for i, c in enumerate(cols) if c == "_status"), None)

    for ci, cn in enumerate(cols, 1):
        cell = ws.cell(1, ci, cn)
        cell.fill, cell.font, cell.border = hdr_fill, hdr_font, border
        cell.alignment = Alignment(horizontal="center")

    for ri, (_, row) in enumerate(df_export.iterrows(), 2):
        status = str(row.iloc[si]) if si is not None and pd.notna(row.iloc[si]) else ""
        if "Cross" in status:           rf, rfo = fills["cross"]
        elif "Consecutive" in status:   rf, rfo = fills["consec"]
        elif "Sum Match" in status:     rf, rfo = fills["sum"]
        elif "Matched" in status and "Unmatched" not in status: rf, rfo = fills["matched"]
        elif "Unmatched" in status:     rf, rfo = fills["unmatched"]
        else:                           rf, rfo = None, default_font
        for ci, cn in enumerate(cols, 1):
            try:
                v = row[cn]; cell = ws.cell(ri, ci, str(v) if pd.notna(v) else "")
            except Exception:
                cell = ws.cell(ri, ci, "")
            cell.font, cell.border = rfo, border
            if rf:
                cell.fill = rf

    for ci, cn in enumerate(cols, 1):
        mx = len(str(cn))
        for r in range(2, min(52, ws.max_row + 1)):
            cv = ws.cell(r, ci).value
            if cv:
                mx = max(mx, len(str(cv)))
        ws.column_dimensions[ws.cell(1, ci).column_letter].width = min(mx + 3, 40)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _color_row(row):
    """Styler helper — colour a dataframe row by its _status."""
    s = str(row.get("_status", "")) if "_status" in row.index else ""
    if "Cross" in s:
        c = "background-color:#fef3c7;color:#92400e"
    elif "Consecutive" in s:
        c = "background-color:#dbeafe;color:#1e40af"
    elif "Sum Match" in s:
        c = "background-color:#ede9fe;color:#5b21b6"
    elif "Matched" in s and "Unmatched" not in s:
        c = "background-color:#dcfce7;color:#166534"
    elif "Unmatched" in s:
        c = "background-color:#fef2f2;color:#991b1b"
    else:
        c = ""
    return [c] * len(row)


# ── Dialogs (Streamlit >= 1.33) ───────────────────────────

@st.dialog("📝 Edit Row", width="large")
def _dlg_edit_row(idx):
    df = st.session_state.df_original
    row = df.loc[idx]
    cols = [c for c in df.columns if not c.startswith("_")]
    vals = {}
    for c in cols:
        cur = str(row[c]) if pd.notna(row[c]) else ""
        vals[c] = st.text_input(c, value=cur, key=f"er_{c}_{idx}")
    if st.button("💾 Save", type="primary", use_container_width=True):
        for c, v in vals.items():
            if v.strip() == "":
                st.session_state.df_original.at[idx, c] = np.nan
            else:
                try:
                    if pd.api.types.is_float_dtype(df[c].dtype):
                        st.session_state.df_original.at[idx, c] = float(v)
                    elif pd.api.types.is_integer_dtype(df[c].dtype):
                        st.session_state.df_original.at[idx, c] = int(float(v))
                    else:
                        st.session_state.df_original.at[idx, c] = v
                except (ValueError, TypeError):
                    st.session_state.df_original.at[idx, c] = v
        st.session_state.df_display = st.session_state.df_original.copy()
        st.rerun()


@st.dialog("➕ Add Row")
def _dlg_add_row():
    df = st.session_state.df_original
    cols = [c for c in df.columns if not c.startswith("_")]
    vals = {c: st.text_input(c, key=f"ar_{c}") for c in cols}
    if st.button("💾 Save Row", type="primary", use_container_width=True):
        new = pd.DataFrame([vals])
        st.session_state.df_original = pd.concat(
            [st.session_state.df_original, new], ignore_index=True
        )
        st.session_state.df_display = st.session_state.df_original.copy()
        st.rerun()


@st.dialog("⚡ Bulk Edit", width="large")
def _dlg_bulk_edit():
    df = st.session_state.df_display
    vis = [c for c in df.columns if not c.startswith("_")]
    col = st.selectbox("Column:", vis)
    val = st.text_input("New value:")
    scope = st.radio(
        "Apply to:",
        [f"Filtered rows ({len(df):,})", f"All rows ({len(st.session_state.df_original):,})"],
        horizontal=True,
    )
    if st.button("✅ Apply", type="primary", use_container_width=True):
        if "Filtered" in scope:
            st.session_state.df_original.loc[df.index, col] = val
        else:
            st.session_state.df_original[col] = val
        st.session_state.df_display = st.session_state.df_original.copy()
        st.rerun()


@st.dialog("➕ Add Computed Column", width="large")
def _dlg_add_column():
    cname = st.text_input("Column name:")
    code = st.text_area(
        "Python code:",
        height=150,
        help="Reference your data as `df`. For multi-line code assign the result to `result`.",
    )
    st.caption(
        "Example: `pd.to_datetime(df['Memo'].str[:10], format='%m/%d/%Y', errors='coerce')`"
    )
    if st.button("✅ Apply", type="primary", use_container_width=True):
        if not cname or not code:
            st.error("Provide both a name and code.")
            return
        try:
            df = st.session_state.df_original
            ns = {
                "df": df, "pd": pd, "np": np, "result": None,
                "str": str, "int": int, "float": float, "len": len,
                "round": round, "abs": abs, "min": min, "max": max,
                "list": list, "dict": dict, "range": range,
                "enumerate": enumerate, "zip": zip, "sorted": sorted,
                "map": map, "filter": filter, "any": any, "all": all,
                "sum": sum, "True": True, "False": False, "None": None,
                "__builtins__": {},
            }
            try:
                res = eval(code, ns, ns)
            except SyntaxError:
                ns2 = dict(ns)
                exec(code, ns2, ns2)
                res = ns2.get("result")
                if res is None:
                    st.error("Multi-line code must assign to `result`.")
                    return
            if not hasattr(res, "__len__") or isinstance(res, str):
                res = pd.Series([res] * len(df), index=df.index)
            st.session_state.df_original[cname] = res
            st.session_state.df_display = st.session_state.df_original.copy()
            st.rerun()
        except Exception as e:
            st.error(f"Error: {e}")


@st.dialog("🔗 Match Details", width="large")
def _dlg_show_match(mid):
    df = st.session_state.df_original
    mdf = df[df["_match_id"] == mid].copy()
    if mdf.empty:
        st.warning(f"No rows with Match ID: {mid}")
        return
    st.subheader(f"Match ID: {mid}")
    if "_match_note" in mdf.columns:
        notes = mdf["_match_note"].dropna().unique()
        if len(notes):
            st.info(str(notes[0]))
    st.caption(f"{len(mdf)} entries in this match group")

    COL_D = st.session_state._col_debit
    COL_C = st.session_state._col_credit
    show = [c for c in mdf.columns if c not in ("_matched", "_match_note")]
    st.dataframe(mdf[show], use_container_width=True, hide_index=True)

    dt = pd.to_numeric(mdf[COL_D], errors="coerce").sum() if COL_D in mdf.columns else 0
    ct = pd.to_numeric(mdf[COL_C], errors="coerce").sum() if COL_C in mdf.columns else 0
    c1, c2, c3 = st.columns(3)
    c1.metric("Debit Total", f"₹{dt:,.2f}")
    c2.metric("Credit Total", f"₹{ct:,.2f}")
    c3.metric("Net", f"₹{dt - ct:,.2f}")


@st.dialog("🗑 Delete Columns", width="large")
def _dlg_delete_columns():
    all_cols = list(st.session_state.df_original.columns)
    to_del = st.multiselect("Select columns to permanently delete:", all_cols)
    st.warning("⚠ This cannot be undone!")
    if st.button("🗑 Delete Selected", type="primary", use_container_width=True):
        if not to_del:
            st.error("Select at least one column.")
            return
        for c in to_del:
            if c in st.session_state.df_original.columns:
                st.session_state.df_original.drop(columns=[c], inplace=True)
            if c in st.session_state.df_display.columns:
                st.session_state.df_display.drop(columns=[c], inplace=True)
            st.session_state.hidden_columns.discard(c)
        st.rerun()


@st.dialog("👁 Hide / Show Columns", width="large")
def _dlg_hide_show():
    all_cols = list(st.session_state.df_original.columns)
    visible = [c for c in all_cols if c not in st.session_state.hidden_columns]
    chosen = st.multiselect(
        "Visible columns (uncheck to hide):", all_cols, default=visible
    )
    if st.button("✅ Apply", type="primary", use_container_width=True):
        st.session_state.hidden_columns = set(all_cols) - set(chosen)
        st.rerun()


# ══════════════════════════════════════════════════════════
#  SIDEBAR
# ══════════════════════════════════════════════════════════

with st.sidebar:
    # ── Load File ─────────────────────────────────────────
    st.markdown("### 📁 Load File")
    st.caption("Upload Excel or CSV")
    uploaded = st.file_uploader(
        "file", type=["xlsx", "xls", "csv"], label_visibility="collapsed"
    )
    if uploaded is not None and uploaded.name != st.session_state.get("_last_upload", ""):
        st.session_state._last_upload = uploaded.name
        st.session_state.file_name = uploaded.name
        ext = os.path.splitext(uploaded.name)[1].lower()
        try:
            raw = uploaded.read()
            st.session_state._uploaded_bytes = raw
            if ext in (".xlsx", ".xls"):
                preview = pd.read_excel(io.BytesIO(raw), header=None, nrows=20, dtype=str)
            else:
                preview = pd.read_csv(
                    io.BytesIO(raw), header=None, nrows=20, dtype=str, encoding="utf-8-sig"
                )
            st.session_state.preview_df = preview
            st.session_state.step = "header_select"
            st.session_state.selected_header = 0
            st.rerun()
        except Exception as e:
            st.error(f"Load error: {e}")

    st.divider()

    # ── Actions ───────────────────────────────────────────
    st.markdown("### ⚙️ Actions")
    a1, a2 = st.columns(2)
    with a1:
        recon_clicked = st.button(
            "📊 Reconcile", use_container_width=True,
            disabled=(st.session_state.df_original is None),
        )
    with a2:
        reset_clicked = st.button(
            "🔄 Reset View", use_container_width=True,
            disabled=(st.session_state.df_original is None),
        )

    st.divider()

    # ── Export ────────────────────────────────────────────
    st.markdown("### 📤 Export")
    if st.session_state.df_display is not None:
        efmt = st.radio("Format", ["Excel", "CSV"], horizontal=True, label_visibility="collapsed")
        df_exp = st.session_state.df_display[
            [c for c in st.session_state.df_display.columns if c != "_matched"]
        ]
        if efmt == "Excel":
            xb = _export_styled_excel(df_exp)
            st.download_button(
                "⬇️ Download Excel", xb, "reconciliation_export.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            cb = df_exp.to_csv(index=False).encode("utf-8")
            st.download_button(
                "⬇️ Download CSV", cb, "reconciliation_export.csv",
                "text/csv", use_container_width=True,
            )
    else:
        st.caption("Load a file first.")

    st.divider()

    # ── Edit Data ─────────────────────────────────────────
    st.markdown("### 🔧 Edit Data")
    if st.session_state.df_original is not None:
        e1, e2 = st.columns(2)
        with e1:
            if st.button("➕ Add Row", use_container_width=True):
                _dlg_add_row()
        with e2:
            if st.button("⚡ Bulk Edit", use_container_width=True):
                _dlg_bulk_edit()

        st.divider()

        st.markdown("### 📊 Columns")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("➕ Column", use_container_width=True):
                _dlg_add_column()
        with c2:
            if st.button("👁 Hide/Show", use_container_width=True):
                _dlg_hide_show()
        if st.button("🗑 Del Columns", use_container_width=True):
            _dlg_delete_columns()
    else:
        st.caption("Load a file first.")


# ══════════════════════════════════════════════════════════
#  MAIN AREA  —  Handle top-level actions first
# ══════════════════════════════════════════════════════════

# Sync globals
COL_DATE = st.session_state._col_date
COL_DEBIT = st.session_state._col_debit
COL_CREDIT = st.session_state._col_credit
COL_OP_DATA = st.session_state._col_op_data
COL_NAME = st.session_state._col_name
COL_EXP_TYPE = st.session_state._col_exp_type

if recon_clicked and st.session_state.df_original is not None:
    with st.spinner("⏳ Running reconciliation — this may take a minute for large files …"):
        result = run_reconciliation(st.session_state.df_original)
        st.session_state.df_original = result
        st.session_state.df_display = result.copy()
        st.session_state.current_page = 0
    st.toast("✅ Reconciliation complete!", icon="🎉")
    st.rerun()

if reset_clicked and st.session_state.df_original is not None:
    st.session_state.df_display = st.session_state.df_original.copy()
    st.session_state.current_page = 0
    st.rerun()


# ── Header ────────────────────────────────────────────────
hdr1, hdr2 = st.columns([3, 1])
with hdr1:
    st.markdown(
        "### ◈ **AccountSync** &nbsp; "
        "<span style='color:#64748b;font-size:0.9rem'>Reconciliation Tool</span>",
        unsafe_allow_html=True,
    )
with hdr2:
    if st.session_state.file_name:
        st.caption(f"📄 {st.session_state.file_name}")


# ══════════════════════════════════════════════════════════
#  STEP: Upload
# ══════════════════════════════════════════════════════════

if st.session_state.step == "upload":
    st.divider()
    _, mid, _ = st.columns([1, 2, 1])
    with mid:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### 📊 Upload a file to get started")
        st.caption("Supports Excel (.xlsx / .xls) and CSV files")
        st.caption("Use the **file uploader** in the sidebar ←")


# ══════════════════════════════════════════════════════════
#  STEP: Header Row Selection
# ══════════════════════════════════════════════════════════

elif st.session_state.step == "header_select":
    st.divider()
    st.subheader("📋 Select Header Row")
    st.caption("Review the first 20 rows and choose which row contains your column headers.")

    preview = st.session_state.preview_df
    if preview is not None:
        disp = preview.copy()
        disp.insert(0, "Row #", range(len(disp)))
        st.dataframe(disp, use_container_width=True, hide_index=True, height=420)

        sel = st.number_input(
            "Header row number:", min_value=0, max_value=len(preview) - 1, value=0
        )

        b1, b2 = st.columns(2)
        with b1:
            if st.button(f"🚀 Use Row {sel} as Header", type="primary", use_container_width=True):
                try:
                    data = st.session_state._uploaded_bytes
                    ext = os.path.splitext(st.session_state.file_name)[1].lower()
                    if ext in (".xlsx", ".xls"):
                        df = pd.read_excel(io.BytesIO(data), header=int(sel), dtype=str)
                    else:
                        df = pd.read_csv(
                            io.BytesIO(data), header=int(sel), dtype=str, encoding="utf-8-sig"
                        )
                    st.session_state.raw_df = df
                    st.session_state.step = "column_map"
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")
        with b2:
            if st.button("❌ Cancel", use_container_width=True):
                st.session_state.step = "upload"
                st.rerun()


# ══════════════════════════════════════════════════════════
#  STEP: Column Mapping
# ══════════════════════════════════════════════════════════

elif st.session_state.step == "column_map":
    st.divider()
    st.subheader("🗂 Column Mapping")
    st.caption("Map each required field to the correct column in your file.  The dropdowns are pre-filled with the best guess.")

    df = st.session_state.raw_df
    file_cols = list(df.columns)
    opts = ["(None — skip)"] + file_cols

    mapping_fields = [
        ("Date Column",         COL_DATE,    ["date", "post date", "posting date", "txn date", "transaction date"]),
        ("Debit Column",        COL_DEBIT,   ["debit", "dr", "debit amount"]),
        ("Credit Column",       COL_CREDIT,  ["credit", "cr", "credit amount"]),
        ("All OP Data Column",  COL_OP_DATA, ["all op data", "op data", "party"]),
        ("Name Column",         COL_NAME,    ["name", "driver name", "employee", "driver"]),
        ("Expense Type Column", COL_EXP_TYPE,["expense type", "type", "category", "expense"]),
    ]

    with st.form("col_map_form"):
        sels = {}
        for label, default, hints in mapping_fields:
            best = _best_match(file_cols, hints, default)
            idx = opts.index(best) if best in opts else 0
            lc, rc = st.columns([1, 2])
            with lc:
                st.markdown(f"**{label}**")
            with rc:
                sels[label] = st.selectbox(
                    label, opts, index=idx, label_visibility="collapsed", key=f"m_{label}"
                )

        fb1, fb2 = st.columns(2)
        with fb1:
            confirm = st.form_submit_button("✅ Confirm Mapping", type="primary")
        with fb2:
            cancel = st.form_submit_button("❌ Cancel")

    if confirm:
        def _g(lab, deflt):
            v = sels.get(lab, "(None — skip)")
            return v if v != "(None — skip)" else deflt

        st.session_state._col_date = _g("Date Column", COL_DATE)
        st.session_state._col_debit = _g("Debit Column", COL_DEBIT)
        st.session_state._col_credit = _g("Credit Column", COL_CREDIT)
        st.session_state._col_op_data = _g("All OP Data Column", COL_OP_DATA)
        st.session_state._col_name = _g("Name Column", COL_NAME)
        st.session_state._col_exp_type = _g("Expense Type Column", COL_EXP_TYPE)

        # Update module-level variables so engine functions see them
        COL_DATE = st.session_state._col_date
        COL_DEBIT = st.session_state._col_debit
        COL_CREDIT = st.session_state._col_credit
        COL_OP_DATA = st.session_state._col_op_data
        COL_NAME = st.session_state._col_name
        COL_EXP_TYPE = st.session_state._col_exp_type

        df = df.copy()
        for col in [COL_DEBIT, COL_CREDIT]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        numeric_cols = {COL_DEBIT, COL_CREDIT}
        for col in df.columns:
            if col in numeric_cols:
                continue
            try:
                nm = df[col].isna()
                s = df[col].fillna("").astype(str).str.strip().str.lower()
                df[col] = s.where(~nm, other=np.nan)
            except Exception:
                pass
        st.session_state.df_original = df
        st.session_state.df_display = df.copy()
        st.session_state.step = "ready"
        st.session_state.current_page = 0
        st.rerun()

    if cancel:
        st.session_state.step = "upload"
        st.rerun()


# ══════════════════════════════════════════════════════════
#  STEP: Ready — Main Data View
# ══════════════════════════════════════════════════════════

elif st.session_state.step == "ready":
    df = st.session_state.df_display
    if df is None:
        st.session_state.step = "upload"
        st.rerun()

    all_cols = list(st.session_state.df_original.columns)

    # ── Filter / Status / Exclude Rectangle ───────────────
    with st.container(border=True):
        filt_col, stat_col, excl_col = st.columns([3, 2, 1])

        with filt_col:
            st.markdown("**FILTER**")
            for i in range(3):
                fc, vc, bc = st.columns([1.5, 2, 0.8])
                with fc:
                    st.selectbox(
                        f"F{i+1} col", all_cols, key=f"fcol_{i}", label_visibility="collapsed"
                    )
                with vc:
                    st.text_input(
                        f"F{i+1} val", key=f"fval_{i}",
                        placeholder="search…", label_visibility="collapsed",
                    )
                with bc:
                    st.checkbox("Blanks", key=f"fblank_{i}")

            ab, cb = st.columns(2)
            with ab:
                if st.button("🔍 Apply", use_container_width=True, key="apply_f"):
                    _apply_filters()
                    st.rerun()
            with cb:
                if st.button("🧹 Clear", use_container_width=True, key="clear_f"):
                    for i in range(3):
                        st.session_state[f"fval_{i}"] = ""
                        st.session_state[f"fblank_{i}"] = False
                    st.session_state.exclude_fuel = False
                    st.session_state.exclude_insurance = False
                    for k in [
                        "✅ Matched", "📊 Consecutive Match", "🧮 Sum Match",
                        "🔄 Cross Matched", "❌ Unmatched",
                    ]:
                        st.session_state[f"st_{k}"] = True
                    st.session_state.df_display = st.session_state.df_original.copy()
                    st.session_state.current_page = 0
                    st.rerun()

        with stat_col:
            st.markdown("**STATUS**")
            for lbl in [
                "✅ Matched", "📊 Consecutive Match", "🧮 Sum Match",
                "🔄 Cross Matched", "❌ Unmatched",
            ]:
                st.checkbox(lbl, value=True, key=f"st_{lbl}")

        with excl_col:
            st.markdown("**EXCLUDE**")
            st.checkbox("⛽ Fuel", key="exclude_fuel")
            st.checkbox("🛡 Insurance", key="exclude_insurance")

    # ── Stats Ribbon ──────────────────────────────────────
    total = len(df)
    if "_status" in df.columns:
        n_exact = int((df["_status"] == "✅ Matched").sum())
        n_con   = int((df["_status"] == "📊 Consecutive Match").sum())
        n_sum   = int((df["_status"] == "🧮 Sum Match").sum())
        n_mat   = n_exact + n_con + n_sum
        n_cross = int((df["_status"] == "🔄 Cross Matched").sum())
        n_unm   = int((df["_status"] == "❌ Unmatched").sum())
    else:
        n_mat = n_con = n_sum = n_cross = n_unm = "—"

    d_sum = pd.to_numeric(df[COL_DEBIT],  errors="coerce").sum() if COL_DEBIT  in df.columns else 0
    c_sum = pd.to_numeric(df[COL_CREDIT], errors="coerce").sum() if COL_CREDIT in df.columns else 0

    m = st.columns(8)
    m[0].metric("ROWS",         f"{total:,}")
    m[1].metric("MATCHED",      f"{n_mat:,}"   if isinstance(n_mat, int) else n_mat)
    m[2].metric("CONSECUTIVE",  f"{n_con:,}"   if isinstance(n_con, int) else n_con)
    m[3].metric("SUM MATCH",    f"{n_sum:,}"   if isinstance(n_sum, int) else n_sum)
    m[4].metric("CROSS",        f"{n_cross:,}" if isinstance(n_cross, int) else n_cross)
    m[5].metric("UNMATCHED",    f"{n_unm:,}"   if isinstance(n_unm, int) else n_unm)
    m[6].metric("TOTAL DEBIT",  f"₹{d_sum:,.0f}")
    m[7].metric("TOTAL CREDIT", f"₹{c_sum:,.0f}")

    # ── Data Table ────────────────────────────────────────
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    page = st.session_state.current_page
    start_idx = page * PAGE_SIZE
    end_idx   = min(start_idx + PAGE_SIZE, total)

    if total > 0:
        page_df = df.iloc[start_idx:end_idx].copy()
        vis_cols = [
            c for c in page_df.columns
            if c not in st.session_state.hidden_columns
            and c not in ("_matched", "_match_note")
        ]
        show_df = page_df[vis_cols]

        styled = show_df.style.apply(_color_row, axis=1)
        event = st.dataframe(
            styled,
            use_container_width=True,
            hide_index=True,
            height=520,
            on_select="rerun",
            selection_mode="multi-row",
        )

        # ── Selection actions ─────────────────────────────
        sel_rows = event.selection.rows if event and event.selection else []
        if sel_rows:
            real_idxs = [
                df.index[start_idx + r] for r in sel_rows if start_idx + r < len(df)
            ]
            sc1, sc2, sc3 = st.columns([1, 1, 3])
            with sc1:
                if len(sel_rows) == 1 and st.button("✏️ Edit Row"):
                    _dlg_edit_row(real_idxs[0])
            with sc2:
                if st.button(f"🗑 Delete {len(sel_rows)} row(s)"):
                    st.session_state.df_original.drop(index=real_idxs, inplace=True)
                    st.session_state.df_original.reset_index(drop=True, inplace=True)
                    st.session_state.df_display = st.session_state.df_original.copy()
                    st.rerun()
            with sc3:
                if "_match_id" in df.columns and len(sel_rows) == 1:
                    mid = str(
                        df.iloc[start_idx + sel_rows[0]].get("_match_id", "")
                    ).strip()
                    if mid:
                        if st.button(f"🔗 Show Match: {mid}"):
                            _dlg_show_match(mid)
    else:
        st.info("No rows match the current filters.")

    # ── Pagination ────────────────────────────────────────
    st.divider()
    p1, p2, p3, p4, p5, p6, p7 = st.columns([0.6, 0.6, 0.6, 0.6, 2, 1.2, 0.8])
    with p1:
        if st.button("«", key="pg_first", disabled=page == 0):
            st.session_state.current_page = 0; st.rerun()
    with p2:
        if st.button("‹", key="pg_prev", disabled=page == 0):
            st.session_state.current_page -= 1; st.rerun()
    with p3:
        if st.button("›", key="pg_next", disabled=page >= total_pages - 1):
            st.session_state.current_page += 1; st.rerun()
    with p4:
        if st.button("»", key="pg_last", disabled=page >= total_pages - 1):
            st.session_state.current_page = total_pages - 1; st.rerun()
    with p5:
        st.markdown(f"**Page {page + 1} / {total_pages}** &nbsp; ({total:,} rows)")
    with p6:
        jump = st.number_input(
            "Go to", 1, total_pages, page + 1, label_visibility="collapsed", key="pg_jump"
        )
    with p7:
        if st.button("Go", key="pg_go"):
            st.session_state.current_page = jump - 1; st.rerun()

